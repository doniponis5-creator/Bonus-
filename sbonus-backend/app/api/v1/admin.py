"""
Sbonus+ — Админ-панель API.
GET  /api/v1/admin/dashboard/stats
POST /api/v1/admin/tiers
POST /api/v1/admin/promo-codes
GET  /api/v1/admin/reports/export
POST /api/v1/admin/cashiers
GET  /api/v1/admin/audit-logs
"""

import io
import uuid
from datetime import datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import UserRole, hash_password, require_role, get_current_user, verify_password
from app.models import (
    AuditLog,
    BonusAccount,
    Coupon,
    Customer,
    PromoCode,
    ReviewRequest,
    ReviewStatus,
    Tier,
    Transaction,
    TransactionType,
    User,
    UserRoleEnum,
)
from app.schemas import (
    DashboardStatsResponse,
    PromoCodeCreateRequest,
    CashierCreateRequest,
    SuccessResponse,
    TierCreateRequest,
    SettingsUpdateRequest,
    AdminCustomerUpdateRequest,
    AdminCashierUpdateRequest,
    AdminBonusAdjustmentRequest,
    BonusResult,
)
from app.models import Setting, CustomerAuthToken
from app.services.whatsapp import send_whatsapp_message
from app.services.bonus import BonusService
from app.services.audit import log_audit
from app.api.v1.wheel import DEFAULT_SEGMENTS

router = APIRouter(prefix="/admin", tags=["Админ-панель"])


def _get_ip(request: Request) -> str:
    """Получить IP клиента (учитывая X-Forwarded-For за Nginx)."""
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _uid(current_user: dict) -> uuid.UUID:
    """Извлечь UUID пользователя из JWT."""
    return uuid.UUID(current_user["sub"])


# ═══════════════════════════════════════════
# PIN VERIFICATION (2FA)
# ═══════════════════════════════════════════

class VerifyPinRequest(BaseModel):
    pin: str


@router.post(
    "/verify-pin",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def verify_admin_pin(
    body: VerifyPinRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> dict:
    """Проверка PIN-кода администратора для критических действий."""
    user = (await db.execute(
        select(User).where(User.id == current_user["sub"])
    )).scalar_one_or_none()

    if not user:
        raise HTTPException(status_code=404, detail={"message": "Пользователь не найден"})

    # Check password_hash (admin uses password, not pin)
    if user.password_hash and verify_password(body.pin, user.password_hash):
        return {"verified": True, "message": "PIN подтверждён"}

    # Also check pin_hash if set
    if user.pin_hash and verify_password(body.pin, user.pin_hash):
        return {"verified": True, "message": "PIN подтверждён"}

    raise HTTPException(status_code=403, detail={"message": "Неверный PIN-код"})


@router.get(
    "/integration/1c-status",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def get_1c_status(
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Статус 1C интеграции: последние операции, ошибки, синхронизация."""
    from app.models import CustomerDebt

    # Check if 1C webhook enabled
    enabled_setting = (await db.execute(
        select(Setting).where(Setting.key == "ENABLE_1C_WEBHOOK")
    )).scalar_one_or_none()
    is_enabled = enabled_setting and enabled_setting.value == "true"

    # Last 1C transaction (with receipt_number)
    last_txn = (await db.execute(
        select(Transaction)
        .where(Transaction.receipt_number.isnot(None))
        .order_by(Transaction.created_at.desc())
        .limit(1)
    )).scalar_one_or_none()

    # Count 1C transactions today
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    txn_today = (await db.execute(
        select(func.count(Transaction.id))
        .where(Transaction.receipt_number.isnot(None), Transaction.created_at >= today_start)
    )).scalar() or 0

    # Last debt sync
    last_debt = (await db.execute(
        select(CustomerDebt).order_by(CustomerDebt.created_at.desc()).limit(1)
    )).scalar_one_or_none()

    # Total debts
    total_debts = (await db.execute(
        select(func.count(CustomerDebt.id))
    )).scalar() or 0

    return {
        "webhook_enabled": is_enabled,
        "last_transaction_at": last_txn.created_at.isoformat() if last_txn else None,
        "last_receipt": last_txn.receipt_number if last_txn else None,
        "transactions_today": txn_today,
        "last_debt_sync_at": last_debt.created_at.isoformat() if last_debt else None,
        "total_debt_records": total_debts,
    }


@router.get(
    "/dashboard/stats",
    response_model=DashboardStatsResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def dashboard_stats(db: AsyncSession = Depends(get_db)) -> DashboardStatsResponse:
    """Общая статистика дашборда — оптимизированная версия."""
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Single query for all customer + bonus stats
    stats_q = await db.execute(
        select(
            func.count(Customer.id).label("total"),
            func.count(Customer.id).filter(Customer.is_active == True).label("active"),
            func.coalesce(func.sum(BonusAccount.total_earned), 0).label("earned"),
            func.coalesce(func.sum(BonusAccount.total_spent), 0).label("spent"),
            func.coalesce(func.sum(BonusAccount.balance), 0).label("balance"),
        ).outerjoin(BonusAccount, Customer.id == BonusAccount.customer_id)
    )
    row = stats_q.one()

    # Single query for transaction counts
    txn_q = await db.execute(
        select(
            func.count(Transaction.id).filter(Transaction.created_at >= today_start).label("today"),
            func.count(Transaction.id).filter(Transaction.created_at >= month_start).label("month"),
        )
    )
    txn_row = txn_q.one()

    # Распределение по уровням
    tier_dist_q = await db.execute(
        select(Tier.name, func.count(Customer.id))
        .outerjoin(Customer, Customer.tier_id == Tier.id)
        .group_by(Tier.name)
    )
    tier_distribution = {name: count for name, count in tier_dist_q.all()}

    return DashboardStatsResponse(
        total_customers=row.total,
        active_customers=row.active,
        total_bonus_issued=Decimal(str(row.earned)),
        total_bonus_spent=Decimal(str(row.spent)),
        total_balance=Decimal(str(row.balance)),
        transactions_today=txn_row.today,
        transactions_month=txn_row.month,
        tier_distribution=tier_distribution,
    )


@router.get(
    "/dashboard/trends",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def dashboard_trends(
    days: int = Query(30, ge=7, le=365),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Тренды для графиков: ежедневная статистика за указанный период.
    Возвращает массивы для recharts: earn/spend/customers по дням.
    """
    from app.models import TransactionType
    since = datetime.now(timezone.utc) - timedelta(days=days)

    # Ежедневные транзакции по типам
    daily_txn = await db.execute(
        select(
            func.date_trunc("day", Transaction.created_at).label("day"),
            Transaction.type,
            func.count(Transaction.id).label("count"),
            func.coalesce(func.sum(Transaction.amount), 0).label("total"),
        )
        .where(Transaction.created_at >= since)
        .group_by("day", Transaction.type)
        .order_by("day")
    )
    txn_rows = daily_txn.all()

    # Ежедневные новые клиенты
    daily_cust = await db.execute(
        select(
            func.date_trunc("day", Customer.created_at).label("day"),
            func.count(Customer.id).label("count"),
        )
        .where(Customer.created_at >= since)
        .group_by("day")
        .order_by("day")
    )
    cust_rows = daily_cust.all()

    # Собираем в dict по дням
    days_map: dict = {}
    for row in txn_rows:
        d = row.day.strftime("%Y-%m-%d")
        if d not in days_map:
            days_map[d] = {"date": d, "earn": 0, "spend": 0, "earn_count": 0, "spend_count": 0, "new_customers": 0}
        if row.type == TransactionType.EARN:
            days_map[d]["earn"] = float(row.total)
            days_map[d]["earn_count"] = row.count
        elif row.type == TransactionType.SPEND:
            days_map[d]["spend"] = float(row.total)
            days_map[d]["spend_count"] = row.count

    for row in cust_rows:
        d = row.day.strftime("%Y-%m-%d")
        if d not in days_map:
            days_map[d] = {"date": d, "earn": 0, "spend": 0, "earn_count": 0, "spend_count": 0, "new_customers": 0}
        days_map[d]["new_customers"] = row.count

    # Топ-5 клиентов по покупкам за период
    top_customers = await db.execute(
        select(
            Customer.full_name,
            Customer.phone,
            func.coalesce(func.sum(Transaction.purchase_amount), 0).label("total_purchase"),
            func.count(Transaction.id).label("txn_count"),
        )
        .join(Transaction, Transaction.customer_id == Customer.id)
        .where(
            Transaction.created_at >= since,
            Transaction.type == TransactionType.EARN,
        )
        .group_by(Customer.id, Customer.full_name, Customer.phone)
        .having(func.coalesce(func.sum(Transaction.purchase_amount), 0) > 0)
        .order_by(func.coalesce(func.sum(Transaction.purchase_amount), 0).desc())
        .limit(5)
    )
    top_rows = top_customers.all()

    # Средний чек за период
    avg_check = await db.execute(
        select(func.avg(Transaction.purchase_amount))
        .where(
            Transaction.created_at >= since,
            Transaction.type == TransactionType.EARN,
            Transaction.purchase_amount.isnot(None),
        )
    )
    avg_val = avg_check.scalar()

    return {
        "daily": sorted(days_map.values(), key=lambda x: x["date"]),
        "top_customers": [
            {
                "name": r.full_name,
                "phone": r.phone,
                "total_purchase": float(r.total_purchase),
                "transactions": r.txn_count,
            }
            for r in top_rows
        ],
        "average_check": round(float(avg_val), 2) if avg_val else 0,
        "period_days": days,
    }


@router.get(
    "/dashboard/notifications",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def dashboard_notifications(
    days: int = Query(7, ge=1, le=90),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Статистика уведомлений: отправлено, failed, retry."""
    from app.models import Notification, NotificationStatus

    since = datetime.now(timezone.utc) - timedelta(days=days)

    stats = await db.execute(
        select(
            Notification.status,
            func.count(Notification.id).label("count"),
        )
        .where(Notification.created_at >= since)
        .group_by(Notification.status)
    )
    status_map = {row.status: row.count for row in stats.all()}

    return {
        "sent": status_map.get("sent", 0),
        "failed": status_map.get("failed", 0),
        "pending": status_map.get("pending", 0),
        "total": sum(status_map.values()),
        "period_days": days,
    }


@router.get(
    "/dashboard/analytics",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def dashboard_analytics(
    days: int = Query(30, ge=7, le=365),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Детальная аналитика: по часам, retention, конверсия, сравнение периодов."""
    since = datetime.now(timezone.utc) - timedelta(days=days)
    prev_since = since - timedelta(days=days)

    # Выручка по часам (для текущего периода)
    hourly = await db.execute(
        select(
            func.extract("hour", Transaction.created_at).label("hour"),
            func.count(Transaction.id).label("count"),
            func.coalesce(func.sum(Transaction.purchase_amount), 0).label("revenue"),
        )
        .where(
            Transaction.created_at >= since,
            Transaction.type == TransactionType.EARN,
        )
        .group_by("hour")
        .order_by("hour")
    )
    hourly_data = [
        {"hour": int(r.hour), "count": r.count, "revenue": float(r.revenue)}
        for r in hourly.all()
    ]

    # Текущий vs предыдущий период
    curr_rev = (await db.execute(
        select(func.coalesce(func.sum(Transaction.purchase_amount), 0))
        .where(Transaction.created_at >= since, Transaction.type == TransactionType.EARN)
    )).scalar() or 0
    prev_rev = (await db.execute(
        select(func.coalesce(func.sum(Transaction.purchase_amount), 0))
        .where(
            Transaction.created_at >= prev_since,
            Transaction.created_at < since,
            Transaction.type == TransactionType.EARN,
        )
    )).scalar() or 0

    curr_customers = (await db.execute(
        select(func.count(Customer.id)).where(Customer.created_at >= since)
    )).scalar() or 0
    prev_customers = (await db.execute(
        select(func.count(Customer.id)).where(
            Customer.created_at >= prev_since, Customer.created_at < since
        )
    )).scalar() or 0

    # Retention: клиенты с >1 покупкой за период
    active_buyers = (await db.execute(
        select(func.count())
        .select_from(
            select(Transaction.customer_id)
            .where(Transaction.created_at >= since, Transaction.type == TransactionType.EARN)
            .group_by(Transaction.customer_id)
            .having(func.count(Transaction.id) > 1)
            .subquery()
        )
    )).scalar() or 0
    total_buyers = (await db.execute(
        select(func.count(func.distinct(Transaction.customer_id)))
        .where(Transaction.created_at >= since, Transaction.type == TransactionType.EARN)
    )).scalar() or 0

    # Распределение по типам транзакций
    type_dist = await db.execute(
        select(
            Transaction.type,
            func.count(Transaction.id).label("count"),
            func.coalesce(func.sum(Transaction.amount), 0).label("total"),
        )
        .where(Transaction.created_at >= since)
        .group_by(Transaction.type)
    )
    type_data = [
        {"type": r.type.value, "count": r.count, "total": float(r.total)}
        for r in type_dist.all()
    ]

    # Средний LTV (total_earned на клиента)
    avg_ltv = (await db.execute(
        select(func.avg(BonusAccount.total_earned))
    )).scalar() or 0

    return {
        "hourly_activity": hourly_data,
        "revenue_current": float(curr_rev),
        "revenue_previous": float(prev_rev),
        "revenue_change_pct": round((float(curr_rev) - float(prev_rev)) / float(prev_rev) * 100, 1) if prev_rev else 0,
        "new_customers_current": curr_customers,
        "new_customers_previous": prev_customers,
        "retention_rate": round(active_buyers / total_buyers * 100, 1) if total_buyers else 0,
        "repeat_buyers": active_buyers,
        "total_buyers": total_buyers,
        "transaction_types": type_data,
        "average_ltv": round(float(avg_ltv), 2),
        "period_days": days,
    }


@router.get(
    "/dashboard/inactive-customers",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def inactive_customers(db: AsyncSession = Depends(get_db)) -> dict:
    """Спящие клиенты — кто давно не покупал, с разбивкой по периодам."""
    now = datetime.now(timezone.utc)

    # Последняя покупка каждого клиента
    last_txn_sub = (
        select(
            Transaction.customer_id,
            func.max(Transaction.created_at).label("last_purchase"),
        )
        .where(Transaction.type == TransactionType.EARN)
        .group_by(Transaction.customer_id)
        .subquery()
    )

    # Все активные клиенты с балансом и последней покупкой
    query = (
        select(
            Customer.id,
            Customer.full_name,
            Customer.phone,
            Customer.created_at,
            BonusAccount.balance,
            last_txn_sub.c.last_purchase,
        )
        .outerjoin(BonusAccount, BonusAccount.customer_id == Customer.id)
        .outerjoin(last_txn_sub, last_txn_sub.c.customer_id == Customer.id)
        .where(Customer.is_active == True)
    )
    result = await db.execute(query)
    rows = result.all()

    # Разбивка по периодам
    buckets = {
        "7_days": {"label": "7 дней", "count": 0, "total_balance": 0, "customers": []},
        "14_days": {"label": "14 дней", "count": 0, "total_balance": 0, "customers": []},
        "30_days": {"label": "30 дней", "count": 0, "total_balance": 0, "customers": []},
        "60_days": {"label": "60 дней", "count": 0, "total_balance": 0, "customers": []},
        "90_days": {"label": "90+ дней", "count": 0, "total_balance": 0, "customers": []},
        "never": {"label": "Ни разу не покупал", "count": 0, "total_balance": 0, "customers": []},
    }

    total_active = len(rows)
    total_sleeping = 0

    for row in rows:
        cid, name, phone, created_at, balance, last_purchase = row
        bal = float(balance) if balance else 0

        if last_purchase is None:
            bucket_key = "never"
        else:
            # Ensure timezone-aware comparison
            lp = last_purchase if last_purchase.tzinfo else last_purchase.replace(tzinfo=timezone.utc)
            days_ago = (now - lp).days
            if days_ago < 7:
                continue  # Активный — пропускаем
            elif days_ago < 14:
                bucket_key = "7_days"
            elif days_ago < 30:
                bucket_key = "14_days"
            elif days_ago < 60:
                bucket_key = "30_days"
            elif days_ago < 90:
                bucket_key = "60_days"
            else:
                bucket_key = "90_days"

        total_sleeping += 1
        buckets[bucket_key]["count"] += 1
        buckets[bucket_key]["total_balance"] += bal
        # Топ-5 клиентов в каждый бакет (по балансу desc)
        if len(buckets[bucket_key]["customers"]) < 5:
            buckets[bucket_key]["customers"].append({
                "id": str(cid),
                "name": name or "—",
                "phone": phone,
                "balance": bal,
                "last_purchase": last_purchase.isoformat() if last_purchase else None,
            })

    # Сортируем клиентов по балансу внутри каждого бакета
    for b in buckets.values():
        b["customers"].sort(key=lambda x: x["balance"], reverse=True)
        b["total_balance"] = round(b["total_balance"], 2)

    return {
        "total_active": total_active,
        "total_sleeping": total_sleeping,
        "sleeping_pct": round(total_sleeping / total_active * 100, 1) if total_active else 0,
        "buckets": buckets,
    }


@router.get(
    "/tiers",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_tiers(db: AsyncSession = Depends(get_db)):
    """Список всех уровней бонусной программы."""
    result = await db.execute(select(Tier).where(Tier.is_active == True).order_by(Tier.sort_order.asc()))
    tiers = result.scalars().all()
    return [
        {
            "id": str(t.id),
            "name": t.name,
            "min_total_kgs": float(t.min_total_kgs),
            "bonus_percent": float(t.bonus_percent),
            "max_spend_pct": float(t.max_spend_pct),
            "sort_order": t.sort_order,
        }
        for t in tiers
    ]


@router.post(
    "/tiers",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def create_or_update_tier(
    body: TierCreateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Добавить или обновить уровень бонусной программы."""
    result = await db.execute(select(Tier).where(Tier.name == body.name))
    existing = result.scalar_one_or_none()

    if existing:
        existing.min_total_kgs = body.min_total_kgs
        existing.bonus_percent = body.bonus_percent
        existing.max_spend_pct = body.max_spend_pct
        await log_audit(db, "tier_update", "tier", existing.id, _uid(current_user),
                        {"name": body.name, "bonus_percent": float(body.bonus_percent)}, _get_ip(request))
        return SuccessResponse(message=f"Уровень '{body.name}' обновлён")
    else:
        max_order = (await db.execute(select(func.max(Tier.sort_order)))).scalar() or 0
        tier = Tier(
            name=body.name,
            min_total_kgs=body.min_total_kgs,
            bonus_percent=body.bonus_percent,
            max_spend_pct=body.max_spend_pct,
            sort_order=max_order + 1,
        )
        db.add(tier)
        await db.flush()
        await log_audit(db, "tier_create", "tier", tier.id, _uid(current_user),
                        {"name": body.name, "bonus_percent": float(body.bonus_percent)}, _get_ip(request))
        return SuccessResponse(message=f"Уровень '{body.name}' создан")


@router.get(
    "/promo-codes",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_promo_codes(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Список промокодов с пагинацией."""
    total = (await db.execute(select(func.count()).select_from(PromoCode))).scalar() or 0
    result = await db.execute(
        select(PromoCode).order_by(PromoCode.created_at.desc()).offset((page - 1) * limit).limit(limit)
    )
    promos = result.scalars().all()
    return {
        "items": [
            {
                "id": str(p.id),
                "code": p.code,
                "bonus_amount": float(p.bonus_amount),
                "max_uses": p.max_uses,
                "used_count": p.used_count,
                "expires_at": p.expires_at.isoformat() if p.expires_at else None,
                "is_active": p.is_active,
                "created_at": p.created_at.isoformat(),
            }
            for p in promos
        ],
        "total": total,
        "page": page,
        "limit": limit,
    }


@router.post(
    "/promo-codes",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def create_promo_code(
    body: PromoCodeCreateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Создать промокод (сумма, срок, лимит)."""
    existing = await db.execute(select(PromoCode).where(PromoCode.code == body.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail={"code": "PROMO_CODE_EXISTS", "message": "Промокод уже существует"})

    promo = PromoCode(
        code=body.code.upper(),
        bonus_amount=body.bonus_amount,
        max_uses=body.max_uses,
        expires_at=body.expires_at,
    )
    db.add(promo)
    await db.flush()
    await log_audit(db, "promo_create", "promo_code", promo.id, _uid(current_user),
                    {"code": body.code.upper(), "bonus": float(body.bonus_amount)}, _get_ip(request))
    return SuccessResponse(message=f"Промокод '{body.code}' создан: +{body.bonus_amount} KGS")


@router.get(
    "/branches",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_branches(db: AsyncSession = Depends(get_db)):
    """Получить список филиалов."""
    from app.models import Branch
    result = await db.execute(select(Branch).order_by(Branch.created_at.asc()))
    branches = result.scalars().all()
    return [
        {
            "id": str(b.id),
            "name": b.name,
            "address": b.address,
            "city": b.city,
            "phone": b.phone,
            "is_active": b.is_active,
            "created_at": b.created_at.isoformat(),
        }
        for b in branches
    ]


class BranchCreateRequest(BaseModel):
    name: str
    address: str | None = None
    city: str | None = None
    phone: str | None = None


@router.post(
    "/branches",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def create_branch(
    body: BranchCreateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Добавить новый филиал."""
    from app.models import Branch
    branch = Branch(name=body.name, address=body.address, city=body.city, phone=body.phone)
    db.add(branch)
    await db.flush()
    await log_audit(db, "branch_create", "branch", branch.id, _uid(current_user),
                    {"name": body.name}, _get_ip(request))
    return SuccessResponse(message=f"Филиал '{body.name}' добавлен")


@router.get(
    "/cashiers",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_cashiers(db: AsyncSession = Depends(get_db)):
    """Список всех кассиров."""
    from app.models import Branch
    result = await db.execute(
        select(User, Branch)
        .outerjoin(Branch, User.branch_id == Branch.id)
        .where(User.role == UserRoleEnum.CASHIER)
        .order_by(User.created_at.desc())
    )
    rows = result.all()
    return [
        {
            "id": str(u.id),
            "full_name": u.full_name,
            "phone": u.phone,
            "branch_id": str(u.branch_id) if u.branch_id else None,
            "branch_name": b.name if b else "—",
            "is_active": u.is_active,
            "created_at": u.created_at.isoformat(),
        }
        for u, b in rows
    ]


@router.post(
    "/cashiers",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def create_cashier(
    body: CashierCreateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Добавить кассира (имя, телефон, PIN, филиал)."""
    existing = await db.execute(select(User).where(User.phone == body.phone))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail={"message": "Кассир с таким телефоном уже существует"})

    cashier = User(
        phone=body.phone,
        full_name=body.full_name,
        role=UserRoleEnum.CASHIER,
        branch_id=body.branch_id,
        pin_hash=hash_password(body.pin),
    )
    db.add(cashier)
    await db.flush()
    await log_audit(db, "cashier_create", "user", cashier.id, _uid(current_user),
                    {"name": body.full_name, "phone": body.phone}, _get_ip(request))
    await db.commit()
    return SuccessResponse(message=f"Кассир '{body.full_name}' добавлен")


@router.patch(
    "/cashiers/{cashier_id}",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def update_cashier(
    cashier_id: uuid.UUID,
    body: AdminCashierUpdateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Обновить кассира (блокировка, переименование, сброс PIN, перевод в другой филиал)."""
    result = await db.execute(
        select(User).where(User.id == cashier_id, User.role == UserRoleEnum.CASHIER)
    )
    cashier = result.scalar_one_or_none()
    if not cashier:
        raise HTTPException(status_code=404, detail={"message": "Кассир не найден"})

    changes = {}
    if body.full_name is not None:
        cashier.full_name = body.full_name
        changes["full_name"] = body.full_name
    if body.branch_id is not None:
        cashier.branch_id = body.branch_id
        changes["branch_id"] = str(body.branch_id)
    if body.is_active is not None:
        cashier.is_active = body.is_active
        changes["is_active"] = body.is_active
    if body.pin is not None:
        cashier.pin_hash = hash_password(body.pin)
        changes["pin_reset"] = True

    await log_audit(db, "cashier_update", "user", cashier_id, _uid(current_user), changes, _get_ip(request))
    await db.commit()
    return SuccessResponse(message="Данные кассира обновлены")


@router.get(
    "/reports/export",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def export_report(
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    days: int = Query(30, ge=1, le=365),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Скачать отчёт по транзакциям в CSV (стриминг) или Excel."""
    since = datetime.now(timezone.utc) - timedelta(days=days)

    if format == "csv":
        async def csv_generator():
            yield "id,customer_id,type,amount,purchase_amount,receipt_number,created_at\n"
            # Use server-side cursor with streaming to avoid loading all into memory
            result = await db.stream(
                select(Transaction)
                .where(Transaction.created_at >= since)
                .order_by(Transaction.created_at.desc())
            )
            async for t in result.scalars():
                yield (
                    f"{t.id},{t.customer_id},{t.type.value},{t.amount},"
                    f"{t.purchase_amount or ''},{t.receipt_number or ''},{t.created_at.isoformat()}\n"
                )
        return StreamingResponse(
            csv_generator(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=sbonus_report_{days}d.csv"},
        )
    else:
        # Excel export — limited to 10 000 rows to prevent OOM
        import openpyxl
        result = await db.execute(
            select(Transaction)
            .where(Transaction.created_at >= since)
            .order_by(Transaction.created_at.desc())
            .limit(10000)
        )
        txns = result.scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакции"
        ws.append(["ID", "Клиент", "Тип", "Сумма", "Покупка", "Чек", "Дата"])
        for t in txns:
            ws.append([str(t.id), str(t.customer_id), t.type.value,
                       float(t.amount), float(t.purchase_amount) if t.purchase_amount else None,
                       t.receipt_number, t.created_at.isoformat()])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sbonus_report_{days}d.xlsx"},
        )


@router.get(
    "/customers",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_customers(
    search: str = Query("", max_length=50),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    tier_name: str = Query(None, description="Фильтр по уровню: Bronze/Silver/Gold/Platinum"),
    is_active: bool = Query(None, description="Фильтр по статусу: true/false"),
    min_balance: float = Query(None, ge=0, description="Минимальный баланс"),
    max_balance: float = Query(None, ge=0, description="Максимальный баланс"),
    sort_by: str = Query("created_at", description="Сортировка: created_at/balance/full_name"),
    sort_dir: str = Query("desc", description="Направление: asc/desc"),
    db: AsyncSession = Depends(get_db),
):
    """Список клиентов с сегментацией, пагинацией и фильтрами."""
    from sqlalchemy import func as sqlfunc
    from app.models import BonusAccount, Customer, Tier

    # Base query with joins
    stmt = (
        select(Customer, BonusAccount, Tier)
        .outerjoin(BonusAccount, Customer.id == BonusAccount.customer_id)
        .outerjoin(Tier, Customer.tier_id == Tier.id)
    )

    # Filters
    if search:
        search_term = f"%{search}%"
        stmt = stmt.where(
            (Customer.phone.ilike(search_term)) |
            (Customer.full_name.ilike(search_term))
        )
    if tier_name:
        stmt = stmt.where(Tier.name == tier_name)
    if is_active is not None:
        stmt = stmt.where(Customer.is_active == is_active)
    if min_balance is not None:
        stmt = stmt.where(BonusAccount.balance >= Decimal(str(min_balance)))
    if max_balance is not None:
        stmt = stmt.where(BonusAccount.balance <= Decimal(str(max_balance)))

    # Count
    count_sub = stmt.subquery()
    total = (await db.execute(select(sqlfunc.count()).select_from(count_sub))).scalar() or 0

    # Sort
    sort_map = {
        "created_at": Customer.created_at,
        "balance": BonusAccount.balance,
        "full_name": Customer.full_name,
    }
    sort_col = sort_map.get(sort_by, Customer.created_at)
    stmt = stmt.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())
    stmt = stmt.offset((page - 1) * limit).limit(limit)

    result = await db.execute(stmt)
    rows = result.all()

    items = []
    for customer, account, tier in rows:
        items.append({
            "id": str(customer.id),
            "full_name": customer.full_name,
            "phone": customer.phone,
            "tier_name": tier.name if tier else "Bronze",
            "balance": float(account.balance) if account else 0.0,
            "total_earned": float(account.total_earned) if account else 0.0,
            "total_spent": float(account.total_spent) if account else 0.0,
            "is_active": customer.is_active,
            "created_at": customer.created_at.isoformat(),
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "limit": limit,
    }


class BulkBonusRequest(BaseModel):
    customer_ids: list[str]
    type: str  # "earn" or "spend"
    amount: Decimal
    note: str


@router.post(
    "/customers/bulk-bonus",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def bulk_bonus(
    body: BulkBonusRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Массовое начисление/списание бонусов для выбранных клиентов."""
    from app.models import TransactionType

    if body.type not in ("earn", "spend"):
        raise HTTPException(status_code=400, detail={"message": "Тип должен быть earn или spend"})
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail={"message": "Сумма должна быть больше 0"})
    if not body.customer_ids:
        raise HTTPException(status_code=400, detail={"message": "Выберите клиентов"})

    tx_type = TransactionType.EARN if body.type == "earn" else TransactionType.SPEND
    admin_id = uuid.UUID(current_user["sub"])
    svc = BonusService(db)

    success = 0
    errors = []
    for cid in body.customer_ids:
        try:
            await svc.admin_adjustment(
                customer_id=uuid.UUID(cid),
                type=tx_type,
                amount=body.amount,
                admin_id=admin_id,
                note=f"[BULK] {body.note}",
            )
            success += 1
        except Exception as e:
            errors.append(f"{cid}: {str(e)[:100]}")

    await log_audit(db, "bulk_bonus", "customer", None, _uid(current_user),
                    {"type": body.type, "amount": body.amount, "count": len(body.customer_ids), "success": success}, _get_ip(request))
    await db.commit()
    msg = f"Обработано: {success}/{len(body.customer_ids)}"
    if errors:
        msg += f". Ошибки: {len(errors)}"
    return SuccessResponse(message=msg)

@router.put(
    "/customers/{id}",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def update_customer(
    id: uuid.UUID,
    body: AdminCustomerUpdateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Обновить данные клиента."""
    from app.models import Customer
    stmt = select(Customer).where(Customer.id == id)
    customer = (await db.execute(stmt)).scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail={"message": "Клиент не найден"})

    # Проверка уникальности телефона
    if body.phone is not None and body.phone != customer.phone:
        existing = (await db.execute(
            select(Customer).where(Customer.phone == body.phone, Customer.id != id)
        )).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=409, detail={"message": f"Телефон {body.phone} уже используется другим клиентом"})
        customer.phone = body.phone

    if body.full_name is not None: customer.full_name = body.full_name
    if body.birth_date is not None: customer.birth_date = body.birth_date
    if body.is_active is not None: customer.is_active = body.is_active

    await log_audit(db, "customer_update", "customer", id, _uid(current_user),
                    body.model_dump(exclude_none=True), _get_ip(request))
    await db.commit()
    return SuccessResponse(message="Данные клиента обновлены")

@router.post(
    "/customers/{id}/bonus/earn",
    response_model=BonusResult,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def earn_admin(
    id: uuid.UUID,
    body: AdminBonusAdjustmentRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Ручное начисление бонуса."""
    from app.models import TransactionType
    svc = BonusService(db)
    res = await svc.admin_adjustment(
        customer_id=id,
        type=TransactionType.EARN,
        amount=body.amount,
        admin_id=uuid.UUID(current_user["sub"]),
        note=body.note
    )
    await log_audit(db, "admin_earn", "customer", id, _uid(current_user),
                    {"amount": float(body.amount), "note": body.note}, _get_ip(request))
    await db.commit()
    return res

@router.post(
    "/customers/{id}/bonus/spend",
    response_model=BonusResult,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def spend_admin(
    id: uuid.UUID,
    body: AdminBonusAdjustmentRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Ручное списание бонуса."""
    from app.models import TransactionType
    svc = BonusService(db)
    res = await svc.admin_adjustment(
        customer_id=id,
        type=TransactionType.SPEND,
        amount=body.amount,
        admin_id=uuid.UUID(current_user["sub"]),
        note=body.note
    )
    await log_audit(db, "admin_spend", "customer", id, _uid(current_user),
                    {"amount": float(body.amount), "note": body.note}, _get_ip(request))
    await db.commit()
    return res

class TransactionReversalRequest(BaseModel):
    reason: str


# ═══════════════════════════════════════════
# GIFT WHEEL SPIN
# ═══════════════════════════════════════════

@router.post(
    "/customers/{id}/gift-spin",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def gift_spin(
    id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Подарить бесплатный спин колеса удачи клиенту."""
    import asyncio
    import secrets
    from datetime import timedelta

    result = await db.execute(select(Customer).where(Customer.id == id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail={"code": "CUSTOMER_NOT_FOUND"})

    # Increment free spins
    free_key = f"WHEEL_FREE_SPINS_{id}"
    result = await db.execute(select(Setting).where(Setting.key == free_key))
    record = result.scalar_one_or_none()
    if record:
        record.value = str(int(record.value) + 1)
    else:
        db.add(Setting(key=free_key, value="1"))

    # Generate magic-link token for direct access
    from app.core.config import get_settings
    cfg = get_settings()
    token_value = secrets.token_urlsafe(32)[:64]
    auth_token = CustomerAuthToken(
        customer_id=id,
        token=token_value,
        expires_at=datetime.now(timezone.utc) + timedelta(days=7),
        ip_address=request.client.host if request.client else None,
    )
    db.add(auth_token)

    cabinet_url = cfg.customer_cabinet_base_url.rstrip("/")
    direct_link = f"{cabinet_url}/wheel?token={token_value}"

    # WhatsApp notification with direct link
    wa_result = await db.execute(
        select(Setting).where(Setting.key.in_([
            "ENABLE_WHATSAPP_NOTIFICATIONS",
            "GREENAPI_INSTANCE_ID",
            "GREENAPI_API_TOKEN",
        ]))
    )
    wa_cfg = {s.key: s.value for s in wa_result.scalars().all()}

    if wa_cfg.get("ENABLE_WHATSAPP_NOTIFICATIONS") == "true":
        instance_id = wa_cfg.get("GREENAPI_INSTANCE_ID")
        api_token = wa_cfg.get("GREENAPI_API_TOKEN")
        if instance_id and api_token:
            msg = (
                f"🎰 Здравствуйте, {customer.full_name}!\n\n"
                f"🎁 Вам подарен бесплатный спин Колеса Удачи!\n"
                f"Испытайте удачу и выиграйте бонусы!\n\n"
                f"👇 Нажмите чтобы крутить:\n{direct_link}"
            )
            asyncio.create_task(send_whatsapp_message(
                phone=customer.phone, message=msg,
                instance_id=instance_id, api_token=api_token
            ))

    await log_audit(db, "gift_spin", "customer", id, _uid(current_user),
                    {"customer_name": customer.full_name}, _get_ip(request))
    await db.commit()
    return SuccessResponse(message=f"Бесплатный спин подарен клиенту {customer.full_name}")


# ═══════════════════════════════════════════
# ДОЛГИ / РАССРОЧКИ КЛИЕНТА
# ═══════════════════════════════════════════

@router.get(
    "/customers/{id}/debts",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_customer_debts(
    id: uuid.UUID,
    status_filter: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> dict:
    """Все долги/рассрочки клиента для админ-панели."""
    from app.models import CustomerDebt

    result = await db.execute(select(Customer).where(Customer.id == id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail={"message": "Клиент не найден"})

    query = select(CustomerDebt).where(CustomerDebt.customer_id == id)
    if status_filter and status_filter in ("active", "overdue", "paid"):
        query = query.where(CustomerDebt.status == status_filter)
    query = query.order_by(CustomerDebt.overdue_days.desc(), CustomerDebt.created_at.desc())

    debts = (await db.execute(query)).scalars().all()

    active_debts = [d for d in debts if d.status != "paid"]
    total_debt = sum(d.amount for d in active_debts)
    total_original = sum(d.total_amount for d in debts)
    total_paid_sum = sum(d.paid_amount for d in debts)
    overdue_count = sum(1 for d in debts if d.overdue_days > 0)

    return {
        "customer_name": customer.full_name,
        "customer_phone": customer.phone,
        "total_debt": float(total_debt),
        "total_original": float(total_original),
        "total_paid": float(total_paid_sum),
        "count": len(debts),
        "overdue_count": overdue_count,
        "debts": [
            {
                "id": str(d.id),
                "reference": d.reference,
                "total_amount": float(d.total_amount),
                "paid_amount": float(d.paid_amount),
                "amount": float(d.amount),
                "overdue_days": d.overdue_days,
                "status": d.status,
                "percent_paid": round(float(d.paid_amount) / float(d.total_amount) * 100, 1) if d.total_amount > 0 else 0,
                "schedule": d.schedule or [],
                "payments_history": d.payments_history or [],
                "next_payment": d.next_payment,
                "note": d.note,
                "created_at": d.created_at.isoformat() if d.created_at else None,
                "synced_at": d.synced_at.isoformat() if d.synced_at else None,
            }
            for d in debts
        ],
    }


@router.get(
    "/debts/summary",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_all_debts_summary(
    status_filter: str | None = None,
    search: str | None = None,
    sort_by: str = "overdue_days",
    limit: int = 50,
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> dict:
    """Сводка по всем должникам для админ-панели (таблица + фильтры)."""
    from app.models import CustomerDebt
    from sqlalchemy import func as sa_func

    # Подзапрос: сгруппировать по клиенту
    subq = (
        select(
            CustomerDebt.customer_id,
            sa_func.sum(CustomerDebt.amount).label("total_debt"),
            sa_func.sum(CustomerDebt.total_amount).label("total_original"),
            sa_func.sum(CustomerDebt.paid_amount).label("total_paid"),
            sa_func.count(CustomerDebt.id).label("debt_count"),
            sa_func.max(CustomerDebt.overdue_days).label("max_overdue"),
            sa_func.max(CustomerDebt.synced_at).label("last_sync"),
        )
        .where(CustomerDebt.status != "paid") if status_filter != "paid" else
        select(
            CustomerDebt.customer_id,
            sa_func.sum(CustomerDebt.amount).label("total_debt"),
            sa_func.sum(CustomerDebt.total_amount).label("total_original"),
            sa_func.sum(CustomerDebt.paid_amount).label("total_paid"),
            sa_func.count(CustomerDebt.id).label("debt_count"),
            sa_func.max(CustomerDebt.overdue_days).label("max_overdue"),
            sa_func.max(CustomerDebt.synced_at).label("last_sync"),
        )
        .where(CustomerDebt.status == "paid")
    )

    if status_filter == "overdue":
        subq = subq.where(CustomerDebt.overdue_days > 0)

    subq = subq.group_by(CustomerDebt.customer_id).subquery()

    # Основной запрос с join на Customer
    query = (
        select(
            Customer.id,
            Customer.full_name,
            Customer.phone,
            subq.c.total_debt,
            subq.c.total_original,
            subq.c.total_paid,
            subq.c.debt_count,
            subq.c.max_overdue,
            subq.c.last_sync,
        )
        .join(subq, Customer.id == subq.c.customer_id)
    )

    if search:
        query = query.where(
            Customer.full_name.ilike(f"%{search}%") | Customer.phone.ilike(f"%{search}%")
        )

    # Сортировка
    if sort_by == "overdue_days":
        query = query.order_by(subq.c.max_overdue.desc())
    elif sort_by == "total_debt":
        query = query.order_by(subq.c.total_debt.desc())
    elif sort_by == "name":
        query = query.order_by(Customer.full_name)
    else:
        query = query.order_by(subq.c.max_overdue.desc())

    # Общие метрики
    count_result = await db.execute(
        select(sa_func.count()).select_from(query.subquery())
    )
    total_count = count_result.scalar() or 0

    # Пагинация
    query = query.offset(offset).limit(limit)
    rows = (await db.execute(query)).all()

    # Общая сводка
    stats_result = await db.execute(
        select(
            sa_func.count(sa_func.distinct(CustomerDebt.customer_id)),
            sa_func.sum(CustomerDebt.amount),
            sa_func.sum(CustomerDebt.total_amount),
            sa_func.sum(CustomerDebt.paid_amount),
        ).where(CustomerDebt.status != "paid")
    )
    stats = stats_result.one()

    return {
        "total_customers": stats[0] or 0,
        "total_debt": float(stats[1] or 0),
        "total_original": float(stats[2] or 0),
        "total_paid": float(stats[3] or 0),
        "count": total_count,
        "customers": [
            {
                "id": str(r[0]),
                "full_name": r[1],
                "phone": r[2],
                "total_debt": float(r[3] or 0),
                "total_original": float(r[4] or 0),
                "total_paid": float(r[5] or 0),
                "debt_count": r[6] or 0,
                "max_overdue": r[7] or 0,
                "last_sync": r[8].isoformat() if r[8] else None,
                "percent_paid": round(float(r[5] or 0) / float(r[4] or 1) * 100, 1),
            }
            for r in rows
        ],
    }



@router.post(
    "/transactions/{txn_id}/reverse",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def reverse_transaction(
    txn_id: uuid.UUID,
    body: TransactionReversalRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """
    Отмена транзакции (REFUND). Таблица transactions иммутабельна,
    поэтому создаём обратную REFUND транзакцию.
    Только EARN-type транзакции можно отменять.
    """
    from app.models import TransactionType, BonusAccount

    result = await db.execute(select(Transaction).where(Transaction.id == txn_id))
    txn = result.scalar_one_or_none()
    if not txn:
        raise HTTPException(status_code=404, detail={"message": "Транзакция не найдена"})

    # Проверяем что не REFUND и не SPEND (нельзя вернуть списание)
    if txn.type in (TransactionType.REFUND, TransactionType.EXPIRE):
        raise HTTPException(status_code=400, detail={"message": "Нельзя отменить REFUND/EXPIRE транзакцию"})

    # Проверяем что ещё не отменена
    existing_refund = await db.execute(
        select(Transaction).where(
            Transaction.customer_id == txn.customer_id,
            Transaction.type == TransactionType.REFUND,
            Transaction.note.contains(str(txn_id)),
        )
    )
    if existing_refund.scalar_one_or_none():
        raise HTTPException(status_code=400, detail={"message": "Эта транзакция уже была отменена"})

    # Получаем аккаунт
    acc_result = await db.execute(
        select(BonusAccount)
        .where(BonusAccount.customer_id == txn.customer_id)
        .with_for_update()
    )
    account = acc_result.scalar_one_or_none()
    if not account:
        raise HTTPException(status_code=400, detail={"message": "Бонусный аккаунт не найден"})

    # Для EARN-type: списываем обратно
    is_earn_type = txn.type in (
        TransactionType.EARN, TransactionType.BIRTHDAY,
        TransactionType.REFERRAL, TransactionType.PROMO, TransactionType.CAMPAIGN,
    )

    if is_earn_type:
        if txn.amount > account.balance:
            raise HTTPException(
                status_code=400,
                detail={"message": f"Недостаточно бонусов для отмены. Баланс: {account.balance} KGS, нужно: {txn.amount} KGS"},
            )
        account.balance -= txn.amount
        account.total_earned -= txn.amount
    elif txn.type == TransactionType.SPEND:
        # Возврат списания — бонусы возвращаются
        account.balance += txn.amount
        account.total_spent -= txn.amount

    # Создаём REFUND
    refund = Transaction(
        customer_id=txn.customer_id,
        type=TransactionType.REFUND,
        amount=txn.amount,
        branch_id=txn.branch_id,
        cashier_id=uuid.UUID(current_user["sub"]),
        receipt_number=f"REFUND-{uuid.uuid4().hex[:10].upper()}",
        note=f"↩ Отмена #{str(txn_id)[:8]}... | {body.reason}",
    )
    db.add(refund)
    await log_audit(db, "transaction_reverse", "transaction", txn_id, _uid(current_user),
                    {"amount": float(txn.amount), "type": txn.type.value, "reason": body.reason}, _get_ip(request))
    await db.commit()

    return SuccessResponse(
        message=f"Транзакция отменена. {'Списано' if is_earn_type else 'Возвращено'} {txn.amount} KGS"
    )


@router.put(
    "/transactions/{txn_id}/cashier",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def update_transaction_cashier(
    txn_id: uuid.UUID,
    body: dict,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Изменить кассира транзакции (raw SQL, обходит immutable trigger)."""
    from sqlalchemy import text as sa_text

    cashier_id = body.get("cashier_id")
    if not cashier_id:
        raise HTTPException(400, "cashier_id обязателен")

    # Verify cashier exists
    cashier = await db.execute(select(User).where(User.id == uuid.UUID(cashier_id)))
    cashier_obj = cashier.scalar_one_or_none()
    if not cashier_obj:
        raise HTTPException(404, "Кассир не найден")

    # Verify transaction exists
    txn = await db.execute(select(Transaction).where(Transaction.id == txn_id))
    txn_obj = txn.scalar_one_or_none()
    if not txn_obj:
        raise HTTPException(404, "Транзакция не найдена")

    old_cashier_id = txn_obj.cashier_id

    # Direct SQL update — bypasses ORM, trigger allows cashier_id updates
    await db.execute(
        sa_text("UPDATE transactions SET cashier_id = :cid WHERE id = :tid"),
        {"cid": str(cashier_id), "tid": str(txn_id)},
    )
    await db.commit()

    # Audit log
    await db.execute(
        AuditLog.__table__.insert().values(
            user_id=uuid.UUID(current_user.get("sub")),
            action="update_transaction_cashier",
            entity_type="transaction",
            entity_id=str(txn_id),
            details={
                "old_cashier_id": str(old_cashier_id) if old_cashier_id else None,
                "new_cashier_id": str(cashier_id),
                "new_cashier_name": cashier_obj.full_name,
            },
            ip_address=_get_ip(request),
        )
    )
    await db.commit()

    return SuccessResponse(message=f"Кассир изменён на: {cashier_obj.full_name}")


@router.get(
    "/transactions",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def get_all_transactions(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    tx_type: str = Query(None, description="Фильтр по типу: earn/spend/referral/birthday/promo"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Все транзакции системы с пагинацией и фильтром по типу."""
    from app.models import Branch, User as UserModel
    from sqlalchemy.orm import selectinload

    query = (
        select(Transaction, Customer, UserModel, Branch)
        .outerjoin(Customer, Transaction.customer_id == Customer.id)
        .outerjoin(UserModel, Transaction.cashier_id == UserModel.id)
        .outerjoin(Branch, Transaction.branch_id == Branch.id)
    )
    if tx_type:
        from app.models import TransactionType
        try:
            query = query.where(Transaction.type == TransactionType(tx_type))
        except ValueError:
            pass

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    result = await db.execute(
        query.order_by(Transaction.created_at.desc())
        .offset((page - 1) * per_page).limit(per_page)
    )
    rows = result.all()

    return {
        "items": [
            {
                "id": str(t.id),
                "type": t.type.value,
                "amount": float(t.amount),
                "purchase_amount": float(t.purchase_amount) if t.purchase_amount else None,
                "receipt_number": t.receipt_number,
                "note": t.note,
                "customer_name": c.full_name if c else "—",
                "customer_phone": c.phone if c else "—",
                "cashier_name": u.full_name if u else "—",
                "branch_name": b.name if b else "—",
                "created_at": t.created_at.isoformat(),
            }
            for t, c, u, b in rows
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.get(
    "/audit-logs",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def get_audit_logs(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    action: str = Query(None),
    entity_type: str = Query(None),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Журнал аудита с фильтрами и пагинацией."""
    query = select(AuditLog, User).outerjoin(User, AuditLog.user_id == User.id)
    if action:
        query = query.where(AuditLog.action == action)
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)

    count_q = select(func.count()).select_from(
        select(AuditLog).where(
            *([AuditLog.action == action] if action else []),
            *([AuditLog.entity_type == entity_type] if entity_type else []),
        ).subquery()
    )
    total = (await db.execute(count_q)).scalar() or 0

    result = await db.execute(
        query.order_by(AuditLog.created_at.desc())
        .offset((page - 1) * per_page).limit(per_page)
    )
    rows = result.all()

    return {
        "items": [
            {
                "id": str(l.id), "user_id": str(l.user_id) if l.user_id else None,
                "user_name": u.full_name if u else None,
                "action": l.action, "entity_type": l.entity_type,
                "entity_id": str(l.entity_id) if l.entity_id else None,
                "details": l.details, "ip_address": l.ip_address,
                "created_at": l.created_at.isoformat(),
            }
            for l, u in rows
        ],
        "total": total, "page": page, "per_page": per_page,
    }


@router.get(
    "/settings",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def get_settings(db: AsyncSession = Depends(get_db)):
    """Получить глобальные настройки. Значение 'None' (из бага старых сохранений) → пустая строка."""
    result = await db.execute(select(Setting))
    settings = result.scalars().all()
    return {s.key: ("" if s.value == "None" else (s.value or "")) for s in settings}


@router.post(
    "/settings",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def update_settings(
    body: SettingsUpdateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Обновить глобальные настройки (None — поле пропускается, существующее значение остаётся)."""
    updates = body.model_dump(exclude_none=True)
    changed_keys = list(updates.keys())
    for key, value in updates.items():
        result = await db.execute(select(Setting).where(Setting.key == key))
        setting = result.scalar_one_or_none()
        if setting:
            setting.value = str(value)
        else:
            db.add(Setting(key=key, value=str(value)))

    await log_audit(db, "settings_update", "settings", None, _uid(current_user),
                    {"keys": changed_keys}, _get_ip(request))
    await db.commit()
    return SuccessResponse(message="Настройки успешно сохранены")


@router.post(
    "/settings/test-whatsapp",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def test_whatsapp(
    phone: str = Query(..., description="Номер телефона в международном формате, например 996557100505"),
    db: AsyncSession = Depends(get_db)
):
    """Отправить тестовое сообщение в WhatsApp."""
    # Получаем настройки
    result = await db.execute(select(Setting).where(Setting.key.in_(["GREENAPI_INSTANCE_ID", "GREENAPI_API_TOKEN", "ENABLE_WHATSAPP_NOTIFICATIONS"])))
    settings_dict = {s.key: s.value for s in result.scalars().all()}
    
    if settings_dict.get("ENABLE_WHATSAPP_NOTIFICATIONS") != "true":
        raise HTTPException(status_code=400, detail={"message": "Уведомления WhatsApp отключены в настройках"})
        
    instance_id = settings_dict.get("GREENAPI_INSTANCE_ID")
    api_token = settings_dict.get("GREENAPI_API_TOKEN")
    
    if not instance_id or not api_token:
        raise HTTPException(status_code=400, detail={"message": "Учетные данные Green API не настроены"})
        
    success = await send_whatsapp_message(
        phone=phone,
        message="✅ Тестовое сообщение от S Bonus+!\nИнтеграция работает успешно.",
        instance_id=instance_id,
        api_token=api_token
    )
    
    if success:
        return SuccessResponse(message="Тестовое сообщение отправлено")
    else:
        raise HTTPException(status_code=500, detail={"message": "Не удалось отправить сообщение. Проверьте консоль для деталей."})


# ═══════════════════════════════════════
# SMART COUPONS (Персональные купоны)
# ═══════════════════════════════════════

@router.post(
    "/coupons/auto-coupon/run",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def run_auto_coupon_now() -> dict:
    """Запустить Auto-Coupon Engine вручную (вместо ожидания cron Чт 11:00)."""
    import asyncio
    from app.services.auto_coupon import run_auto_coupon
    asyncio.create_task(run_auto_coupon())
    return {"success": True, "message": "Auto-Coupon Engine запущен в фоновом режиме"}


@router.post(
    "/notifications/post-purchase/run",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def run_post_purchase_now() -> dict:
    """Запустить Post-Purchase Follow-up вручную (вместо ожидания cron 11:10)."""
    import asyncio
    from app.services.smart_notifications import run_post_purchase_followup
    asyncio.create_task(run_post_purchase_followup())
    return {"success": True, "message": "Post-Purchase Follow-up запущен в фоновом режиме"}


class CouponCreateRequest(BaseModel):
    customer_id: str | None = None  # None = доступен всем
    title: str
    description: str | None = None
    bonus_amount: float
    min_purchase: float = 0
    expires_at: str | None = None  # ISO datetime


@router.get(
    "/coupons",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def list_coupons(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    customer_id: str = Query(None),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Список всех купонов с фильтром по клиенту."""
    from sqlalchemy.orm import selectinload

    q = select(Coupon).options(selectinload(Coupon.customer))
    if customer_id:
        q = q.where(Coupon.customer_id == customer_id)

    total = (await db.execute(
        select(func.count()).select_from(q.subquery())
    )).scalar() or 0

    result = await db.execute(
        q.order_by(Coupon.created_at.desc())
        .offset((page - 1) * limit).limit(limit)
    )
    coupons = result.scalars().all()

    return {
        "items": [
            {
                "id": str(c.id),
                "code": c.code,
                "title": c.title,
                "description": c.description,
                "bonus_amount": float(c.bonus_amount),
                "min_purchase": float(c.min_purchase),
                "customer_id": str(c.customer_id) if c.customer_id else None,
                "customer_name": c.customer.full_name if c.customer else "Все клиенты",
                "is_used": c.is_used,
                "is_active": c.is_active,
                "expires_at": c.expires_at.isoformat() if c.expires_at else None,
                "used_at": c.used_at.isoformat() if c.used_at else None,
                "created_at": c.created_at.isoformat(),
            }
            for c in coupons
        ],
        "total": total,
        "page": page,
    }


@router.post(
    "/coupons",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def create_coupon(
    body: CouponCreateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Создать персональный или общий купон."""
    import secrets

    code = f"SC-{secrets.token_hex(4).upper()}"

    expires = None
    if body.expires_at:
        expires = datetime.fromisoformat(body.expires_at)

    coupon = Coupon(
        customer_id=uuid.UUID(body.customer_id) if body.customer_id else None,
        code=code,
        title=body.title,
        description=body.description,
        bonus_amount=Decimal(str(body.bonus_amount)),
        min_purchase=Decimal(str(body.min_purchase)),
        expires_at=expires,
    )
    db.add(coupon)
    await db.flush()
    await log_audit(db, "coupon_create", "coupon", coupon.id, _uid(current_user),
                    {"code": code, "title": body.title, "bonus": float(body.bonus_amount)}, _get_ip(request))
    await db.commit()

    # Отправить WhatsApp если персональный купон
    if body.customer_id:
        customer = (await db.execute(
            select(Customer).where(Customer.id == body.customer_id)
        )).scalar_one_or_none()
        if customer:
            wa_settings = await db.execute(
                select(Setting).where(Setting.key.in_([
                    "ENABLE_WHATSAPP_NOTIFICATIONS", "GREENAPI_INSTANCE_ID", "GREENAPI_API_TOKEN",
                ]))
            )
            s_map = {s.key: s.value for s in wa_settings.scalars().all()}
            if s_map.get("ENABLE_WHATSAPP_NOTIFICATIONS") == "true":
                iid = s_map.get("GREENAPI_INSTANCE_ID")
                tok = s_map.get("GREENAPI_API_TOKEN")
                if iid and tok:
                    msg = (
                        f"🎟 {customer.full_name}, у вас новый купон!\n"
                        f"*{body.title}*\n"
                        f"Бонус: +{int(body.bonus_amount)} KGS\n"
                        f"Код: {code}\n\n"
                        f"📱 Активируйте: https://cabinet.smartcentr.store\n"
                        f"🛒 Смарт Центр"
                    )
                    await send_whatsapp_message(
                        phone=customer.phone, message=msg,
                        instance_id=iid, api_token=tok,
                    )

    return SuccessResponse(message=f"Купон создан: {code}")


@router.delete(
    "/coupons/{coupon_id}",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def delete_coupon(
    coupon_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Деактивировать купон."""
    result = await db.execute(select(Coupon).where(Coupon.id == coupon_id))
    coupon = result.scalar_one_or_none()
    if not coupon:
        raise HTTPException(status_code=404, detail={"message": "Купон не найден"})
    coupon.is_active = False
    await log_audit(db, "coupon_delete", "coupon", coupon_id, _uid(current_user),
                    {"code": coupon.code, "title": coupon.title}, _get_ip(request))
    await db.commit()
    return SuccessResponse(message="Купон деактивирован")


# ═══════════════════════════════════════════
# REVIEW BONUS MANAGEMENT
# ═══════════════════════════════════════════

@router.get(
    "/reviews",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def list_reviews(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    status_filter: str = Query(None, alias="status"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Список заявок на бонус за отзыв."""
    from sqlalchemy.orm import selectinload

    q = select(ReviewRequest).options(selectinload(ReviewRequest.customer))
    if status_filter:
        try:
            q = q.where(ReviewRequest.status == ReviewStatus(status_filter))
        except ValueError:
            pass

    total = (await db.execute(
        select(func.count()).select_from(q.subquery())
    )).scalar() or 0

    result = await db.execute(
        q.order_by(ReviewRequest.created_at.desc())
        .offset((page - 1) * limit).limit(limit)
    )
    reviews = result.scalars().all()

    return {
        "items": [
            {
                "id": str(r.id),
                "customer_id": str(r.customer_id),
                "customer_name": r.customer.full_name if r.customer else "—",
                "customer_phone": r.customer.phone if r.customer else "—",
                "platform": r.platform.value,
                "review_link": r.review_link,
                "status": r.status.value,
                "bonus_amount": float(r.bonus_amount),
                "admin_note": r.admin_note,
                "created_at": r.created_at.isoformat(),
                "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
            }
            for r in reviews
        ],
        "total": total,
        "page": page,
    }


class ReviewActionRequest(BaseModel):
    action: str  # "approve" or "reject"
    note: str | None = None


@router.post(
    "/reviews/{review_id}",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))],
)
async def action_review(
    review_id: uuid.UUID,
    body: ReviewActionRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Одобрить или отклонить заявку на бонус за отзыв."""
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(ReviewRequest).options(selectinload(ReviewRequest.customer))
        .where(ReviewRequest.id == review_id)
    )
    review = result.scalar_one_or_none()
    if not review:
        raise HTTPException(status_code=404, detail={"message": "Заявка не найдена"})
    if review.status != ReviewStatus.PENDING:
        raise HTTPException(status_code=400, detail={"message": "Заявка уже обработана"})

    review.reviewed_by = uuid.UUID(current_user["sub"])
    review.reviewed_at = datetime.now(timezone.utc)
    review.admin_note = body.note

    if body.action == "approve":
        review.status = ReviewStatus.APPROVED

        # Credit bonus
        account = (await db.execute(
            select(BonusAccount).where(BonusAccount.customer_id == review.customer_id).with_for_update()
        )).scalar_one_or_none()
        if account:
            account.balance += review.bonus_amount
            account.total_earned += review.bonus_amount

            platform_name = "Google Maps" if review.platform.value == "google" else "2GIS"
            txn = Transaction(
                customer_id=review.customer_id,
                type=TransactionType.PROMO,
                amount=review.bonus_amount,
                note=f"⭐ Бонус за отзыв на {platform_name}",
            )
            db.add(txn)

            # WhatsApp notification
            wa_settings = await db.execute(
                select(Setting).where(Setting.key.in_([
                    "ENABLE_WHATSAPP_NOTIFICATIONS", "GREENAPI_INSTANCE_ID", "GREENAPI_API_TOKEN",
                ]))
            )
            s_map = {s.key: s.value for s in wa_settings.scalars().all()}
            if s_map.get("ENABLE_WHATSAPP_NOTIFICATIONS") == "true" and review.customer:
                iid = s_map.get("GREENAPI_INSTANCE_ID")
                tok = s_map.get("GREENAPI_API_TOKEN")
                if iid and tok:
                    msg = (
                        f"⭐ {review.customer.full_name}, спасибо за отзыв!\n"
                        f"Вам начислено +{int(review.bonus_amount)} KGS бонусов.\n"
                        f"Баланс: {float(account.balance):.0f} KGS\n\n"
                        f"📱 https://cabinet.smartcentr.store\n"
                        f"🛒 Смарт Центр"
                    )
                    await send_whatsapp_message(
                        phone=review.customer.phone, message=msg,
                        instance_id=iid, api_token=tok,
                    )

        await log_audit(db, "review_approve", "review", review_id, _uid(current_user),
                        {"bonus": float(review.bonus_amount)}, _get_ip(request))
        await db.commit()
        return SuccessResponse(message=f"Отзыв одобрен, +{int(review.bonus_amount)} KGS начислено")

    elif body.action == "reject":
        review.status = ReviewStatus.REJECTED
        await log_audit(db, "review_reject", "review", review_id, _uid(current_user),
                        {"note": body.note}, _get_ip(request))
        await db.commit()
        return SuccessResponse(message="Заявка отклонена")

    else:
        raise HTTPException(status_code=400, detail={"message": "Действие: approve или reject"})


# ═══════════════════════════════════════════
# WHEEL CONFIGURATION (Колесо удачи)
# ═══════════════════════════════════════════

class WheelSegmentInput(BaseModel):
    id: int
    label: str
    value: int  # bonus amount (0 = no prize)
    color: str
    probability: float  # 0.0 - 1.0
    prize_type: str = "bonus"  # "bonus" | "physical" | "none"
    stock: int | None = None  # None = безлимит; иначе макс. выигрышей за период (запас приза)
    stock_period: str = "total"  # "day" | "week" | "month" | "total" — окно квоты


class WheelConfigUpdateRequest(BaseModel):
    segments: list[WheelSegmentInput]


@router.get(
    "/wheel/config",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def get_wheel_config(db: AsyncSession = Depends(get_db)) -> dict:
    """Получить текущую конфигурацию колеса удачи."""
    import json
    result = await db.execute(
        select(Setting).where(Setting.key == "WHEEL_SEGMENTS")
    )
    setting = result.scalar_one_or_none()
    if setting and setting.value:
        try:
            segments = json.loads(setting.value)
            return {"segments": segments, "source": "database"}
        except Exception:
            pass
    return {"segments": DEFAULT_SEGMENTS, "source": "default"}


@router.put(
    "/wheel/config",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def update_wheel_config(
    body: WheelConfigUpdateRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Обновить конфигурацию колеса удачи (сегменты, вероятности, бонусы)."""
    import json

    if len(body.segments) < 2:
        raise HTTPException(status_code=400, detail={"message": "Минимум 2 сегмента"})
    if len(body.segments) > 12:
        raise HTTPException(status_code=400, detail={"message": "Максимум 12 сегментов"})

    # Validate probabilities
    total_prob = sum(s.probability for s in body.segments)
    if abs(total_prob - 1.0) > 0.01:
        raise HTTPException(
            status_code=400,
            detail={"message": f"Сумма вероятностей должна быть 1.0 (сейчас: {total_prob:.4f})"},
        )

    # Validate each segment
    for s in body.segments:
        if s.probability < 0 or s.probability > 1:
            raise HTTPException(status_code=400, detail={"message": f"Вероятность сегмента '{s.label}' вне диапазона 0-1"})
        if s.value < 0:
            raise HTTPException(status_code=400, detail={"message": f"Бонус сегмента '{s.label}' не может быть отрицательным"})
        if not s.label.strip():
            raise HTTPException(status_code=400, detail={"message": "Название сегмента не может быть пустым"})
        if not s.color.strip():
            raise HTTPException(status_code=400, detail={"message": "Цвет сегмента обязателен"})

    # Re-assign sequential IDs
    segments_data = []
    for i, s in enumerate(body.segments, 1):
        seg = {
            "id": i,
            "label": s.label.strip(),
            "value": s.value,
            "color": s.color.strip(),
            "probability": round(s.probability, 4),
            "prize_type": s.prize_type if s.prize_type in ("bonus", "physical", "none") else "bonus",
        }
        # Квота приза (запас): сохраняем только если задан
        if s.stock is not None and int(s.stock) > 0:
            seg["stock"] = int(s.stock)
            seg["stock_period"] = s.stock_period if s.stock_period in ("day", "week", "month", "total") else "total"
        segments_data.append(seg)

    # Save to DB
    result = await db.execute(
        select(Setting).where(Setting.key == "WHEEL_SEGMENTS")
    )
    setting = result.scalar_one_or_none()
    json_value = json.dumps(segments_data, ensure_ascii=False)

    if setting:
        setting.value = json_value
    else:
        db.add(Setting(key="WHEEL_SEGMENTS", value=json_value))

    await log_audit(db, "wheel_config_update", "wheel", None, _uid(current_user),
                    {"segments_count": len(segments_data)}, _get_ip(request))
    await db.commit()
    return SuccessResponse(message=f"Конфигурация колеса сохранена: {len(segments_data)} сегментов")


@router.post(
    "/wheel/config/reset",
    response_model=SuccessResponse,
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def reset_wheel_config(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
) -> SuccessResponse:
    """Сбросить конфигурацию колеса к значениям по умолчанию."""
    import json
    result = await db.execute(
        select(Setting).where(Setting.key == "WHEEL_SEGMENTS")
    )
    setting = result.scalar_one_or_none()
    if setting:
        setting.value = json.dumps(DEFAULT_SEGMENTS, ensure_ascii=False)
    else:
        db.add(Setting(key="WHEEL_SEGMENTS", value=json.dumps(DEFAULT_SEGMENTS, ensure_ascii=False)))

    await log_audit(db, "wheel_config_reset", "wheel", None, _uid(current_user), None, _get_ip(request))
    await db.commit()
    return SuccessResponse(message="Конфигурация колеса сброшена к значениям по умолчанию")


# ─── Excel Bulk Import ──────────────────────────────────────────

@router.post(
    "/customers/import",
    summary="Импорт клиентов из Excel",
    dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))],
)
async def import_customers_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Массовый импорт клиентов из Excel файла.
    Ожидаемые столбцы: ФИО и номер телефона.
    Для каждого клиента создаются QR-код, реферальный код, уровень Bronze и бонусный аккаунт.
    Дубликаты (существующий номер телефона) пропускаются.
    """
    import openpyxl
    from app.utils import normalize_phone

    # Проверка типа файла
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail={"message": "Принимаются только файлы .xlsx или .xls"},
        )

    # Чтение файла
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(
            status_code=400,
            detail={"message": "Не удалось прочитать Excel файл. Проверьте формат."},
        )

    # Уровень по умолчанию (Bronze — минимальный sort_order)
    tier_result = await db.execute(
        select(Tier).order_by(Tier.sort_order.asc()).limit(1)
    )
    default_tier = tier_result.scalar_one_or_none()

    # Предварительная загрузка существующих телефонов (для скорости)
    existing_phones_result = await db.execute(select(Customer.phone))
    existing_phones = {row[0] for row in existing_phones_result.all()}

    # Бесплатные спины для новых клиентов (настройка колеса удачи)
    free_spins_result = await db.execute(
        select(Setting).where(Setting.key == "WHEEL_FREE_SPINS_ON_REGISTER")
    )
    free_spins_record = free_spins_result.scalar_one_or_none()
    free_spins_count = int(free_spins_record.value) if free_spins_record and free_spins_record.value else 0

    created = 0
    skipped = 0
    errors = []
    row_num = 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        row_num += 1

        # Пропуск пустых строк
        if not row or not any(row):
            continue

        # Определение столбцов (2 или более)
        cells = [str(c).strip() if c is not None else "" for c in row]

        # Пропуск строки заголовка
        lower_cells = [c.lower() for c in cells]
        if any(kw in lower_cells for kw in ["фио", "fio", "имя", "ism", "name", "телефон", "phone", "номер"]):
            continue

        # Поиск ФИО и телефона
        fio = ""
        phone_raw = ""

        if len(cells) >= 2:
            # Если первый столбец — порядковый номер, следующий — ФИО
            first_clean = cells[0].replace(" ", "").replace("+", "").replace("-", "")
            if first_clean.isdigit() and len(first_clean) <= 5 and len(cells) >= 3:
                # Порядковый номер | ФИО | Телефон
                fio = cells[1]
                phone_raw = cells[2]
            else:
                # ФИО | Телефон
                fio = cells[0]
                phone_raw = cells[1]

        if not fio or not phone_raw:
            if any(cells):
                errors.append({"row": row_num, "reason": "ФИО или телефон не найден", "data": " | ".join(cells[:3])})
            continue

        # Валидация ФИО
        fio = fio.strip()
        if len(fio) < 2 or len(fio) > 100:
            errors.append({"row": row_num, "reason": f"Неверная длина ФИО: {len(fio)}", "data": fio})
            continue

        # Нормализация телефона
        try:
            phone = normalize_phone(phone_raw)
        except Exception:
            errors.append({"row": row_num, "reason": "Неверный номер телефона", "data": phone_raw})
            continue

        if not phone or len(phone) < 10:
            errors.append({"row": row_num, "reason": "Неверный номер телефона", "data": phone_raw})
            continue

        # Проверка дубликатов
        if phone in existing_phones:
            skipped += 1
            continue

        # Генерация QR и реферального кода
        qr_code = f"SB-{uuid.uuid4().hex[:10].upper()}"
        referral_code = f"REF-{uuid.uuid4().hex[:8].upper()}"

        # Создание клиента
        customer = Customer(
            phone=phone,
            full_name=fio,
            qr_code=qr_code,
            referral_code=referral_code,
            tier_id=default_tier.id if default_tier else None,
            is_active=True,
        )
        db.add(customer)
        await db.flush()  # Получение ID

        # Создание бонусного аккаунта (баланс 0)
        bonus_account = BonusAccount(
            customer_id=customer.id,
            balance=0,
            total_earned=0,
            total_spent=0,
        )
        db.add(bonus_account)

        # Начисление бесплатных спинов колеса удачи
        if free_spins_count > 0:
            db.add(Setting(
                key=f"WHEEL_FREE_SPINS_{customer.id}",
                value=str(free_spins_count),
            ))

        existing_phones.add(phone)
        created += 1

    await log_audit(db, "customer_import", "customer", None, _uid(current_user),
                    {"created": created, "skipped": skipped, "errors": len(errors), "file": file.filename}, _get_ip(request))
    await db.commit()
    wb.close()

    return {
        "message": f"Импорт завершён: добавлено {created} новых клиентов",
        "created": created,
        "skipped": skipped,
        "errors_count": len(errors),
        "errors": errors[:50],  # Первые 50 ошибок
        "total_rows": row_num,
    }


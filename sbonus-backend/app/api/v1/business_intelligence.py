"""
SBonus+ — PRO Бизнес-аналитика.

6 модулей:
  1. Telegram P&L отчёт (cron)
  2. Excel экспорт P&L
  3. Бюджет / лимиты по категориям
  4. Қарздорлар (дебиторская задолженность) аналитика
  5. Кассир KPI scoring
  6. RFM Pro сегментация с рекомендациями

Endpoints:
  GET  /bi/tg-pnl-preview       — превью TG P&L отчёта
  POST /bi/tg-pnl-send          — отправить P&L в Telegram вручную
  GET  /bi/export-excel          — скачать P&L Excel
  GET  /bi/budgets               — лимиты по категориям
  PUT  /bi/budgets               — установить лимит
  GET  /bi/budget-alerts         — алерты превышения
  GET  /bi/debts-analytics       — аналитика задолженностей
  GET  /bi/debts-risk            — рисковые должники
  GET  /bi/cashier-kpi           — KPI скоринг кассиров
  GET  /bi/cashier-kpi/{id}      — детальный KPI кассира
  GET  /bi/rfm-pro               — RFM Pro сегментация
  GET  /bi/rfm-pro/{segment}     — клиенты сегмента
"""

import io

import uuid
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import desc, func, select, and_, case, distinct, literal_column
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db, async_session
from app.core.security import require_role, UserRole
from app.core.logging import get_logger
from app.models import (
    Customer, CustomerDebt,
    Expense, EXPENSE_CATEGORY_LABELS,
    Product, PurchaseItem, Setting,
    Transaction, TransactionType,
    User,
)

logger = get_logger("business_intelligence")

router = APIRouter(
    prefix="/bi",
    tags=["PRO Бизнес-аналитика"],
)


# ═══════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════

def _month_range(month_str: str):
    y, m = int(month_str[:4]), int(month_str[5:7])
    start = datetime(y, m, 1, tzinfo=timezone.utc)
    end = datetime(y + 1, 1, 1, tzinfo=timezone.utc) if m == 12 else datetime(y, m + 1, 1, tzinfo=timezone.utc)
    return start, end

def _current_month() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m")

def _fmt(v) -> str:
    return f"{float(v):,.0f}".replace(",", " ")


# ═══════════════════════════════════════════════════
#  1. TELEGRAM P&L ОТЧЁТ
# ═══════════════════════════════════════════════════

async def _build_pnl_text(db: AsyncSession, month: str) -> str:
    """Строит текстовый P&L отчёт для Telegram."""
    start, end = _month_range(month)

    # Revenue
    rev_q = await db.execute(
        select(func.coalesce(func.sum(PurchaseItem.total), 0))
        .where(PurchaseItem.created_at >= start, PurchaseItem.created_at < end)
    )
    revenue = float(rev_q.scalar() or 0)

    # COGS
    cogs_q = await db.execute(
        select(func.coalesce(func.sum(PurchaseItem.quantity * Product.cost_price), 0))
        .join(Product, Product.id == PurchaseItem.product_id)
        .where(PurchaseItem.created_at >= start, PurchaseItem.created_at < end)
    )
    cogs = float(cogs_q.scalar() or 0)
    gross = revenue - cogs

    # OpEx
    opex_q = await db.execute(
        select(func.coalesce(func.sum(Expense.amount), 0))
        .where(Expense.month == month)
    )
    opex = float(opex_q.scalar() or 0)

    # Bonus expense
    bonus_types = [TransactionType.EARN, TransactionType.BIRTHDAY, TransactionType.REFERRAL,
                   TransactionType.PROMO, TransactionType.CAMPAIGN]
    bonus_q = await db.execute(
        select(func.coalesce(func.sum(Transaction.amount), 0))
        .where(Transaction.type.in_(bonus_types),
               Transaction.created_at >= start, Transaction.created_at < end)
    )
    bonus_exp = float(bonus_q.scalar() or 0)

    net_profit = gross - opex - bonus_exp
    margin = round(net_profit / revenue * 100, 1) if revenue > 0 else 0

    # Top expense categories
    cat_q = await db.execute(
        select(Expense.category, func.sum(Expense.amount).label("total"))
        .where(Expense.month == month)
        .group_by(Expense.category)
        .order_by(desc("total"))
        .limit(5)
    )
    top_cats = cat_q.all()

    # Txn count
    txn_q = await db.execute(
        select(func.count(distinct(Transaction.id)))
        .where(Transaction.type == TransactionType.EARN,
               Transaction.created_at >= start, Transaction.created_at < end)
    )
    txn_count = txn_q.scalar() or 0
    avg_receipt = round(revenue / max(txn_count, 1), 0)

    emoji = "📈" if net_profit > 0 else "📉"
    lines = [
        f"{emoji} <b>P&L отчёт — {month}</b>",
        "",
        f"💰 Выручка: <b>{_fmt(revenue)} сом</b>",
        f"📦 Себестоимость: {_fmt(cogs)} сом",
        f"📊 Валовая прибыль: <b>{_fmt(gross)} сом</b>",
        "",
        f"🏢 Операционные расходы: {_fmt(opex)} сом",
        f"🎁 Бонусные расходы: {_fmt(bonus_exp)} сом",
        "",
        f"{'✅' if net_profit > 0 else '❌'} <b>Чистая прибыль: {_fmt(net_profit)} сом</b>",
        f"📐 Маржинальность: {margin}%",
        "",
        f"🧾 Чеков: {txn_count} | Средний чек: {_fmt(avg_receipt)} сом",
    ]

    if top_cats:
        lines.append("")
        lines.append("📋 <b>ТОП расходы:</b>")
        for cat in top_cats:
            label = EXPENSE_CATEGORY_LABELS.get(cat.category, cat.category)
            lines.append(f"  • {label}: {_fmt(cat.total)} сом")

    return "\n".join(lines)


@router.get("/tg-pnl-preview", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def tg_pnl_preview(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    m = month or _current_month()
    text = await _build_pnl_text(db, m)
    return {"month": m, "preview": text}


@router.post("/tg-pnl-send", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def tg_pnl_send(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    m = month or _current_month()
    text = await _build_pnl_text(db, m)

    # Send via existing telegram bot infrastructure
    from app.services.telegram_bot import _get_bot, _get_chat_id
    bot = await _get_bot(db)
    chat_id = await _get_chat_id(db)
    if not bot or not chat_id:
        raise HTTPException(400, "Telegram бот не настроен")
    await bot.send_message(chat_id, text)
    return {"success": True, "message": "P&L отчёт отправлен в Telegram"}


async def send_pnl_telegram_report():
    """Cron задача — автоматический P&L отчёт в Telegram (21:30)."""
    async with async_session() as db:
        try:
            from app.services.telegram_bot import _get_bot, _get_chat_id
            bot = await _get_bot(db)
            chat_id = await _get_chat_id(db)
            if not bot or not chat_id:
                return
            text = await _build_pnl_text(db, _current_month())
            await bot.send_message(chat_id, text)
            logger.info("Telegram: P&L report sent")
        except Exception as e:
            logger.error(f"Telegram P&L report error: {e}")


# ═══════════════════════════════════════════════════
#  2. EXCEL ЭКСПОРТ P&L
# ═══════════════════════════════════════════════════

@router.get("/export-excel", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def export_excel(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    months: int = Query(3, ge=1, le=12),
    db: AsyncSession = Depends(get_db),
):
    """Скачать P&L отчёт в Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")

    target = month or _current_month()

    wb = openpyxl.Workbook()

    # ─── Sheet 1: P&L Summary ───
    ws = wb.active
    ws.title = "P&L Отчёт"

    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    sub_font = Font(bold=True, size=11, color="FFE600")
    sub_fill = PatternFill(start_color="0d1526", end_color="0d1526", fill_type="solid")
    money_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    border = Border(
        bottom=Side(style='thin', color='333333')
    )

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"P&L Отчёт — Смарт Центр — {target}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    # Build months list
    m_list = []
    y, m = int(target[:4]), int(target[5:7])
    for i in range(months):
        m_list.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    m_list.reverse()

    # Headers
    ws['A3'] = 'Показатель'
    ws['A3'].font = bold_font
    for col_idx, ml in enumerate(m_list, start=2):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = ml
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    rows_def = [
        ("Выручка", "revenue"),
        ("Себестоимость", "cogs"),
        ("Валовая прибыль", "gross"),
        ("", None),
        ("Операционные расходы", "opex"),
        ("Бонусные расходы", "bonus"),
        ("", None),
        ("ЧИСТАЯ ПРИБЫЛЬ", "net"),
        ("Маржинальность %", "margin"),
        ("Количество чеков", "txn_count"),
        ("Средний чек", "avg_receipt"),
    ]

    # Gather data for each month
    month_data = {}
    for ml in m_list:
        s, e = _month_range(ml)

        rev_q = await db.execute(
            select(func.coalesce(func.sum(PurchaseItem.total), 0))
            .where(PurchaseItem.created_at >= s, PurchaseItem.created_at < e)
        )
        revenue = float(rev_q.scalar() or 0)

        cogs_q = await db.execute(
            select(func.coalesce(func.sum(PurchaseItem.quantity * Product.cost_price), 0))
            .join(Product, Product.id == PurchaseItem.product_id)
            .where(PurchaseItem.created_at >= s, PurchaseItem.created_at < e)
        )
        cogs = float(cogs_q.scalar() or 0)

        opex_q = await db.execute(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(Expense.month == ml)
        )
        opex = float(opex_q.scalar() or 0)

        bonus_types = [TransactionType.EARN, TransactionType.BIRTHDAY, TransactionType.REFERRAL,
                       TransactionType.PROMO, TransactionType.CAMPAIGN]
        bonus_q = await db.execute(
            select(func.coalesce(func.sum(Transaction.amount), 0))
            .where(Transaction.type.in_(bonus_types),
                   Transaction.created_at >= s, Transaction.created_at < e)
        )
        bonus = float(bonus_q.scalar() or 0)

        txn_q = await db.execute(
            select(func.count(distinct(Transaction.id)))
            .where(Transaction.type == TransactionType.EARN,
                   Transaction.created_at >= s, Transaction.created_at < e)
        )
        txn_count = txn_q.scalar() or 0

        gross = revenue - cogs
        net = gross - opex - bonus
        margin = round(net / revenue * 100, 1) if revenue > 0 else 0
        avg_receipt = round(revenue / max(txn_count, 1), 0)

        month_data[ml] = {
            "revenue": revenue, "cogs": cogs, "gross": gross,
            "opex": opex, "bonus": bonus, "net": net,
            "margin": margin, "txn_count": txn_count, "avg_receipt": avg_receipt,
        }

    row = 4
    for label, key in rows_def:
        ws.cell(row=row, column=1, value=label).font = bold_font if key in ("net", "gross") else money_font
        ws.cell(row=row, column=1).border = border
        if key:
            for col_idx, ml in enumerate(m_list, start=2):
                val = month_data[ml][key]
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.number_format = '#,##0' if key != 'margin' else '0.0"%"'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                if key == "net":
                    cell.font = Font(bold=True, color="22C55E" if val >= 0 else "EF4444")
        row += 1

    # ─── Sheet 2: Расходы по категориям ───
    ws2 = wb.create_sheet("Расходы по категориям")
    ws2['A1'] = 'Категория'
    ws2['A1'].font = bold_font
    for col_idx, ml in enumerate(m_list, start=2):
        ws2.cell(row=1, column=col_idx, value=ml).font = bold_font

    cats_q = await db.execute(
        select(Expense.category, Expense.month, func.sum(Expense.amount).label("total"))
        .where(Expense.month.in_(m_list))
        .group_by(Expense.category, Expense.month)
    )
    cats_data = defaultdict(lambda: defaultdict(float))
    for r in cats_q.all():
        cats_data[r.category][r.month] = float(r.total)

    row2 = 2
    for cat, months_dict in sorted(cats_data.items()):
        ws2.cell(row=row2, column=1, value=EXPENSE_CATEGORY_LABELS.get(cat, cat))
        for col_idx, ml in enumerate(m_list, start=2):
            ws2.cell(row=row2, column=col_idx, value=months_dict.get(ml, 0)).number_format = '#,##0'
        row2 += 1

    # Column widths
    for sheet in [ws, ws2]:
        sheet.column_dimensions['A'].width = 25
        for i in range(2, len(m_list) + 2):
            sheet.column_dimensions[get_column_letter(i)].width = 16

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PnL_SmartCentr_{target}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════
#  3. БЮДЖЕТ / ЛИМИТЫ
# ═══════════════════════════════════════════════════

class BudgetSet(BaseModel):
    category: str = Field(..., description="Категория расхода")
    limit_amount: float = Field(..., gt=0, description="Лимит в сом")
    month: str = Field(..., pattern=r"^\d{4}-\d{2}$")


@router.get("/budgets", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def get_budgets(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Лимиты по категориям на месяц + факт."""
    m = month or _current_month()
    prefix = f"BUDGET_{m}_"

    # Get all budget limits for month
    limits_q = await db.execute(
        select(Setting).where(Setting.key.like(f"{prefix}%"))
    )
    limits = {s.key.replace(prefix, ""): float(s.value) for s in limits_q.scalars().all()}

    # Get actual expenses
    actual_q = await db.execute(
        select(Expense.category, func.sum(Expense.amount).label("total"))
        .where(Expense.month == m)
        .group_by(Expense.category)
    )
    actuals = {r.category: float(r.total) for r in actual_q.all()}

    budgets = []
    all_cats = set(list(limits.keys()) + list(actuals.keys()))
    for cat in sorted(all_cats):
        limit = limits.get(cat, 0)
        actual = actuals.get(cat, 0)
        pct = round(actual / limit * 100, 1) if limit > 0 else 0
        status = "ok"
        if limit > 0:
            if pct >= 100:
                status = "exceeded"
            elif pct >= 80:
                status = "warning"
        budgets.append({
            "category": cat,
            "label": EXPENSE_CATEGORY_LABELS.get(cat, cat),
            "limit": limit,
            "actual": actual,
            "percent": pct,
            "remaining": max(limit - actual, 0),
            "status": status,
        })

    return {"month": m, "budgets": budgets}


@router.put("/budgets", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def set_budget(
    body: BudgetSet,
    db: AsyncSession = Depends(get_db),
) -> dict:
    key = f"BUDGET_{body.month}_{body.category}"
    result = await db.execute(select(Setting).where(Setting.key == key))
    setting = result.scalar_one_or_none()
    if setting:
        setting.value = str(body.limit_amount)
    else:
        db.add(Setting(key=key, value=str(body.limit_amount)))
    await db.commit()
    return {"success": True, "key": key}


@router.get("/budget-alerts", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def budget_alerts(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Алерты — какие категории превысили/близки к лимиту."""
    data = await get_budgets(month, db)
    alerts = [b for b in data["budgets"] if b["status"] in ("exceeded", "warning")]
    return {"month": data["month"], "alerts": alerts, "count": len(alerts)}


# ═══════════════════════════════════════════════════
#  4. ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ
# ═══════════════════════════════════════════════════

@router.get("/debts-analytics", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))])
async def debts_analytics(
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Полная аналитика дебиторской задолженности."""
    # Total stats
    stats_q = await db.execute(
        select(
            func.count(CustomerDebt.id).label("total_debts"),
            func.coalesce(func.sum(CustomerDebt.total_amount), 0).label("total_amount"),
            func.coalesce(func.sum(CustomerDebt.paid_amount), 0).label("total_paid"),
            func.coalesce(func.sum(CustomerDebt.amount), 0).label("total_remaining"),
            func.count(case((CustomerDebt.status == "overdue", 1))).label("overdue_count"),
            func.coalesce(func.sum(case(
                (CustomerDebt.status == "overdue", CustomerDebt.amount),
                else_=literal_column("0")
            )), 0).label("overdue_amount"),
        ).where(CustomerDebt.status != "paid")
    )
    stats = stats_q.one()

    # By status
    status_q = await db.execute(
        select(
            CustomerDebt.status,
            func.count(CustomerDebt.id).label("count"),
            func.coalesce(func.sum(CustomerDebt.amount), 0).label("amount"),
        ).group_by(CustomerDebt.status)
    )
    by_status = [{"status": r.status, "count": r.count, "amount": float(r.amount)} for r in status_q.all()]

    # Overdue aging buckets: 0-30, 31-60, 61-90, 90+
    aging_q = await db.execute(
        select(
            case(
                (CustomerDebt.overdue_days <= 30, "0-30"),
                (CustomerDebt.overdue_days <= 60, "31-60"),
                (CustomerDebt.overdue_days <= 90, "61-90"),
                else_="90+"
            ).label("bucket"),
            func.count(CustomerDebt.id).label("count"),
            func.coalesce(func.sum(CustomerDebt.amount), 0).label("amount"),
        ).where(CustomerDebt.overdue_days > 0, CustomerDebt.status != "paid")
        .group_by("bucket")
    )
    aging = [{"bucket": r.bucket, "count": r.count, "amount": float(r.amount)} for r in aging_q.all()]

    # Payment rate
    paid_rate = round(float(stats.total_paid) / float(stats.total_amount) * 100, 1) if float(stats.total_amount) > 0 else 0

    return {
        "total_debts": stats.total_debts,
        "total_amount": float(stats.total_amount),
        "total_paid": float(stats.total_paid),
        "total_remaining": float(stats.total_remaining),
        "overdue_count": stats.overdue_count,
        "overdue_amount": float(stats.overdue_amount),
        "paid_rate": paid_rate,
        "by_status": by_status,
        "aging": aging,
    }


# ─── Кредит рейтинг калькулятор ───
def _calc_credit_score(debts_data: list[dict]) -> int:
    """Кредит рейтинг 0-100. Юқори = яхши."""
    if not debts_data:
        return 50  # Янги клиент — ўрта

    score = 60  # Базовый

    total_debts = len(debts_data)
    paid_debts = sum(1 for d in debts_data if d["status"] == "paid")
    active_debts = sum(1 for d in debts_data if d["status"] != "paid")
    max_overdue = max((d["overdue_days"] for d in debts_data), default=0)
    total_amount = sum(d["amount"] for d in debts_data if d["status"] != "paid")
    total_ever = sum(d["total_amount"] for d in debts_data)
    total_paid = sum(d["paid_amount"] for d in debts_data)

    # 1. Тўлов тарихи (+20 / -20)
    if total_ever > 0:
        paid_ratio = total_paid / total_ever
        score += int(paid_ratio * 20)  # 0..+20
    else:
        score += 10

    # 2. Просрочка (-35 max)
    if max_overdue > 180:
        score -= 35
    elif max_overdue > 90:
        score -= 28
    elif max_overdue > 60:
        score -= 20
    elif max_overdue > 30:
        score -= 12
    elif max_overdue > 0:
        score -= 5

    # 3. Актив долглар сони (-15 max)
    if active_debts >= 3:
        score -= 15
    elif active_debts == 2:
        score -= 8
    elif active_debts == 1:
        score -= 3

    # 4. Тўланган долглар бонус (+10)
    if paid_debts > 0 and active_debts == 0:
        score += 10
    elif paid_debts > active_debts:
        score += 5

    # 5. Қолдиқ суммаси (-10 max)
    if total_amount > 100000:
        score -= 10
    elif total_amount > 50000:
        score -= 5

    return max(0, min(100, score))


def _auto_category(credit_score: int, debts_data: list[dict]) -> str:
    """Автоматик категория — 5 сегмент."""
    if not debts_data:
        return "new"

    active = [d for d in debts_data if d["status"] != "paid"]
    max_overdue = max((d["overdue_days"] for d in active), default=0)
    overdue_count = sum(1 for d in active if d["overdue_days"] > 30)

    # Чёрный список: 90+ кун ёки 2+ просрочка > 30
    if max_overdue > 90 or overdue_count >= 2:
        return "blacklist"

    # Проблемный: 30-90 кун
    if max_overdue > 30:
        return "problematic"

    # На контроле: 1-30 кун
    if max_overdue > 0:
        return "monitoring"

    # Надёжный: просрочка йўқ
    if len(active) > 0 or any(d["status"] == "paid" for d in debts_data):
        return "reliable"

    return "new"


def _installment_recommendation(category: str, credit_score: int) -> dict:
    """Рассрочка тавсияси."""
    if category == "blacklist":
        return {"allowed": False, "label": "Запрещено", "color": "#ef4444",
                "reason": "Клиент в чёрном списке — хроническая просрочка"}
    if category == "problematic":
        return {"allowed": False, "label": "Не рекомендуется", "color": "#f97316",
                "reason": "Просрочка 30-90 дней — высокий риск невозврата"}
    if category == "monitoring":
        return {"allowed": True, "label": "С осторожностью", "color": "#f59e0b",
                "reason": "Небольшая просрочка — малая сумма, короткий срок"}
    if category == "new":
        return {"allowed": True, "label": "Малая сумма", "color": "#3b82f6",
                "reason": "Новый клиент — начните с небольшой суммы"}
    # reliable
    return {"allowed": True, "label": "Разрешено", "color": "#22c55e",
            "reason": "Надёжный клиент — оплачивает вовремя"}


CATEGORY_META = {
    "blacklist":   {"label": "Чёрный список", "icon": "ban",           "color": "#ef4444", "sort": 0},
    "problematic": {"label": "Проблемный",    "icon": "alert-triangle", "color": "#f97316", "sort": 1},
    "monitoring":  {"label": "На контроле",   "icon": "eye",           "color": "#f59e0b", "sort": 2},
    "reliable":    {"label": "Надёжный",      "icon": "shield-check",  "color": "#22c55e", "sort": 3},
    "new":         {"label": "Новый",         "icon": "user-plus",     "color": "#3b82f6", "sort": 4},
}


@router.get("/debts-registry", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))])
async def debts_registry(
    category: str = Query(None, description="blacklist/problematic/monitoring/reliable/new"),
    search: str = Query(None, description="Поиск по телефону или имени"),
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=10, le=100),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Полный реестр должников с категоризацией и кредитным рейтингом."""
    # 1. Все долги сгруппированные по клиенту
    q = (
        select(
            Customer.id.label("cid"),
            Customer.full_name,
            Customer.phone,
            Customer.is_active,
            func.count(CustomerDebt.id).label("debt_count"),
            func.sum(CustomerDebt.amount).label("total_remaining"),
            func.sum(CustomerDebt.total_amount).label("total_amount"),
            func.sum(CustomerDebt.paid_amount).label("total_paid"),
            func.max(CustomerDebt.overdue_days).label("max_overdue"),
            func.array_agg(CustomerDebt.status).label("statuses"),
            func.array_agg(CustomerDebt.overdue_days).label("overdue_arr"),
            func.array_agg(CustomerDebt.amount).label("amount_arr"),
            func.array_agg(CustomerDebt.total_amount).label("total_arr"),
            func.array_agg(CustomerDebt.paid_amount).label("paid_arr"),
            func.array_agg(CustomerDebt.id).label("debt_ids"),
        )
        .join(Customer, Customer.id == CustomerDebt.customer_id)
        .group_by(Customer.id, Customer.full_name, Customer.phone, Customer.is_active)
    )

    # Поиск
    if search:
        search_term = f"%{search.strip()}%"
        q = q.where(
            (Customer.phone.ilike(search_term)) | (Customer.full_name.ilike(search_term))
        )

    result = await db.execute(q)
    all_rows = result.all()

    # 2. Загрузить админские override-ы
    override_keys = [f"DEBT_CAT_{r.cid}" for r in all_rows]
    overrides: dict[str, str] = {}
    if override_keys:
        ov_q = await db.execute(
            select(Setting.key, Setting.value).where(Setting.key.in_(override_keys))
        )
        for ok in ov_q.all():
            cid = ok.key.replace("DEBT_CAT_", "")
            overrides[cid] = ok.value

    # 3. Вычислить категорию + рейтинг для каждого клиента
    customers = []
    for r in all_rows:
        debts_data = []
        for i in range(len(r.statuses)):
            debts_data.append({
                "status": r.statuses[i],
                "overdue_days": r.overdue_arr[i],
                "amount": float(r.amount_arr[i]),
                "total_amount": float(r.total_arr[i]),
                "paid_amount": float(r.paid_arr[i]),
            })

        credit_score = _calc_credit_score(debts_data)
        auto_cat = _auto_category(credit_score, debts_data)
        cid_str = str(r.cid)
        admin_cat = overrides.get(cid_str)
        final_cat = admin_cat if admin_cat and admin_cat in CATEGORY_META else auto_cat

        recommendation = _installment_recommendation(final_cat, credit_score)

        customers.append({
            "customer_id": cid_str,
            "name": r.full_name or "—",
            "phone": r.phone,
            "is_active": r.is_active,
            "debt_count": r.debt_count,
            "total_remaining": float(r.total_remaining or 0),
            "total_amount": float(r.total_amount or 0),
            "total_paid": float(r.total_paid or 0),
            "max_overdue": r.max_overdue or 0,
            "credit_score": credit_score,
            "auto_category": auto_cat,
            "admin_override": admin_cat,
            "category": final_cat,
            "category_meta": CATEGORY_META[final_cat],
            "recommendation": recommendation,
        })

    # 4. Фильтр по категории
    if category and category in CATEGORY_META:
        customers = [c for c in customers if c["category"] == category]

    # Сортировка: blacklist → new, внутри по max_overdue desc
    customers.sort(key=lambda c: (CATEGORY_META[c["category"]]["sort"], -c["max_overdue"]))

    # 5. Сводка по категориям
    summary = {}
    for cat, meta in CATEGORY_META.items():
        cat_items = [c for c in customers if c["category"] == cat] if not category else                     [c for c in [cc for cc in all_rows] if False]  # placeholder
        summary[cat] = {**meta, "count": 0, "total_debt": 0}

    # Пересчитать summary из полного списка (без фильтра)
    all_customers_cats: dict[str, list] = {k: [] for k in CATEGORY_META}
    for r in all_rows:
        debts_data = []
        for i in range(len(r.statuses)):
            debts_data.append({
                "status": r.statuses[i], "overdue_days": r.overdue_arr[i],
                "amount": float(r.amount_arr[i]), "total_amount": float(r.total_arr[i]),
                "paid_amount": float(r.paid_arr[i]),
            })
        cs = _calc_credit_score(debts_data)
        ac = _auto_category(cs, debts_data)
        cid_str = str(r.cid)
        fc = overrides.get(cid_str) if overrides.get(cid_str) in CATEGORY_META else ac
        all_customers_cats[fc].append(float(r.total_remaining or 0))

    summary = {}
    for cat, meta in CATEGORY_META.items():
        items = all_customers_cats.get(cat, [])
        summary[cat] = {**meta, "count": len(items), "total_debt": round(sum(items), 2)}

    # 6. Пагинация
    total = len(customers)
    start = (page - 1) * per_page
    paginated = customers[start:start + per_page]

    return {
        "customers": paginated,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
        "summary": summary,
        "categories_meta": CATEGORY_META,
    }


@router.put("/debts-override", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN))])
async def debts_override(
    customer_id: str = Query(...),
    new_category: str = Query(...),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Админ вручную устанавливает категорию клиента (override)."""
    if new_category not in CATEGORY_META and new_category != "auto":
        return {"error": "Неверная категория"}

    key = f"DEBT_CAT_{customer_id}"

    if new_category == "auto":
        # Удалить override — вернуться к авто
        existing = await db.execute(select(Setting).where(Setting.key == key))
        setting = existing.scalar_one_or_none()
        if setting:
            await db.delete(setting)
            await db.commit()
        return {"status": "ok", "message": "Категория сброшена на автоматическую"}

    # Upsert
    existing = await db.execute(select(Setting).where(Setting.key == key))
    setting = existing.scalar_one_or_none()
    if setting:
        setting.value = new_category
    else:
        db.add(Setting(key=key, value=new_category))
    await db.commit()

    return {"status": "ok", "category": new_category, "label": CATEGORY_META[new_category]["label"]}


@router.get("/debt-check/{phone}", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN, UserRole.CASHIER))])
async def debt_check(
    phone: str,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Быстрая проверка клиента по телефону перед выдачей рассрочки."""
    # Нормализация
    p = phone.strip().replace(" ", "").replace("-", "")
    if not p.startswith("+"):
        p = "+" + p

    # Найти клиента
    cust_q = await db.execute(
        select(Customer).where(Customer.phone.ilike(f"%{p[-9:]}%"))
    )
    customer = cust_q.scalar_one_or_none()

    if not customer:
        return {
            "found": False,
            "message": "Клиент не найден в базе",
            "recommendation": {"allowed": True, "label": "Новый клиент", "color": "#3b82f6",
                               "reason": "Клиент не в базе — начните с малой суммы"},
        }

    # Долги
    debts_q = await db.execute(
        select(CustomerDebt).where(CustomerDebt.customer_id == customer.id)
    )
    debts = debts_q.scalars().all()

    debts_data = [{
        "status": d.status,
        "overdue_days": d.overdue_days,
        "amount": float(d.amount),
        "total_amount": float(d.total_amount),
        "paid_amount": float(d.paid_amount),
    } for d in debts]

    credit_score = _calc_credit_score(debts_data)
    auto_cat = _auto_category(credit_score, debts_data)

    # Override?
    ov = await db.execute(select(Setting.value).where(Setting.key == f"DEBT_CAT_{customer.id}"))
    admin_cat = ov.scalar_one_or_none()
    final_cat = admin_cat if admin_cat and admin_cat in CATEGORY_META else auto_cat

    recommendation = _installment_recommendation(final_cat, credit_score)

    active_debts = [d for d in debts if d.status != "paid"]
    return {
        "found": True,
        "customer_id": str(customer.id),
        "name": customer.full_name,
        "phone": customer.phone,
        "credit_score": credit_score,
        "category": final_cat,
        "category_meta": CATEGORY_META[final_cat],
        "admin_override": admin_cat,
        "active_debts": len(active_debts),
        "total_remaining": sum(float(d.amount) for d in active_debts),
        "max_overdue": max((d.overdue_days for d in active_debts), default=0),
        "debts": [{
            "id": str(d.id),
            "total_amount": float(d.total_amount),
            "paid_amount": float(d.paid_amount),
            "remaining": float(d.amount),
            "overdue_days": d.overdue_days,
            "status": d.status,
            "reference": d.reference,
        } for d in debts],
        "recommendation": recommendation,
    }


# ═══════════════════════════════════════════════════
#  5. КАССИР KPI
# ═══════════════════════════════════════════════════

@router.get("/cashier-kpi", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))])
async def cashier_kpi(
    month: str = Query(None, pattern=r"^\d{4}-\d{2}$"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """KPI скоринг всех кассиров за месяц."""
    m = month or _current_month()
    start, end = _month_range(m)

    # All cashiers
    cashiers_q = await db.execute(
        select(User).where(User.role == "cashier", User.is_active == True)
    )
    cashiers = cashiers_q.scalars().all()

    if not cashiers:
        return {"month": m, "kpis": [], "team_avg": {}}

    kpis = []
    totals = {"revenue": 0, "txn_count": 0, "customers_served": 0, "bonus_earn": 0, "bonus_spend": 0}

    for c in cashiers:
        # Revenue (sum of purchase_amount from EARN transactions)
        rev_q = await db.execute(
            select(
                func.coalesce(func.sum(Transaction.purchase_amount), 0).label("revenue"),
                func.count(Transaction.id).label("txn_count"),
                func.count(distinct(Transaction.customer_id)).label("unique_customers"),
                func.coalesce(func.avg(Transaction.purchase_amount), 0).label("avg_receipt"),
            )
            .where(
                Transaction.cashier_id == c.id,
                Transaction.type == TransactionType.EARN,
                Transaction.created_at >= start,
                Transaction.created_at < end,
            )
        )
        rev = rev_q.one()

        # Bonus earned (by their customers)
        earn_q = await db.execute(
            select(func.coalesce(func.sum(Transaction.amount), 0))
            .where(
                Transaction.cashier_id == c.id,
                Transaction.type == TransactionType.EARN,
                Transaction.created_at >= start,
                Transaction.created_at < end,
            )
        )
        bonus_earned = float(earn_q.scalar() or 0)

        # Bonus spent (processed by this cashier)
        spend_q = await db.execute(
            select(func.coalesce(func.sum(Transaction.amount), 0))
            .where(
                Transaction.cashier_id == c.id,
                Transaction.type == TransactionType.SPEND,
                Transaction.created_at >= start,
                Transaction.created_at < end,
            )
        )
        bonus_spent = float(spend_q.scalar() or 0)

        # New customers registered by this cashier (from transactions where they appeared first)
        revenue = float(rev.revenue)
        txn_count = rev.txn_count
        unique_cust = rev.unique_customers
        avg_receipt = float(rev.avg_receipt)

        # KPI Score (weighted formula)
        # Revenue: 40%, Txn Count: 20%, Unique Customers: 20%, Avg Receipt: 20%
        # Normalized relative to team max later
        kpis.append({
            "cashier_id": str(c.id),
            "name": c.full_name or "—",
            "branch_id": str(c.branch_id) if c.branch_id else None,
            "revenue": revenue,
            "txn_count": txn_count,
            "unique_customers": unique_cust,
            "avg_receipt": round(avg_receipt, 0),
            "bonus_earned": bonus_earned,
            "bonus_spent": bonus_spent,
            "score": 0,  # will calculate
            "rank": 0,
            "grade": "",
        })
        totals["revenue"] += revenue
        totals["txn_count"] += txn_count
        totals["customers_served"] += unique_cust
        totals["bonus_earn"] += bonus_earned
        totals["bonus_spend"] += bonus_spent

    # Normalize and score
    max_rev = max((k["revenue"] for k in kpis), default=1) or 1
    max_txn = max((k["txn_count"] for k in kpis), default=1) or 1
    max_cust = max((k["unique_customers"] for k in kpis), default=1) or 1
    max_avg = max((k["avg_receipt"] for k in kpis), default=1) or 1

    for k in kpis:
        score = (
            (k["revenue"] / max_rev) * 40 +
            (k["txn_count"] / max_txn) * 20 +
            (k["unique_customers"] / max_cust) * 20 +
            (k["avg_receipt"] / max_avg) * 20
        )
        k["score"] = round(score, 1)

    # Sort by score descending
    kpis.sort(key=lambda x: x["score"], reverse=True)
    for i, k in enumerate(kpis):
        k["rank"] = i + 1
        if k["score"] >= 80:
            k["grade"] = "A+"
        elif k["score"] >= 60:
            k["grade"] = "A"
        elif k["score"] >= 40:
            k["grade"] = "B"
        elif k["score"] >= 20:
            k["grade"] = "C"
        else:
            k["grade"] = "D"

    n = len(kpis) or 1
    team_avg = {
        "avg_revenue": round(totals["revenue"] / n, 0),
        "avg_txn_count": round(totals["txn_count"] / n, 0),
        "avg_customers": round(totals["customers_served"] / n, 0),
        "total_revenue": totals["revenue"],
        "total_txn": totals["txn_count"],
    }

    return {"month": m, "kpis": kpis, "team_avg": team_avg, "cashier_count": n}


# ═══════════════════════════════════════════════════
#  6. RFM PRO СЕГМЕНТАЦИЯ
# ═══════════════════════════════════════════════════

RFM_SEGMENTS = {
    "champions": {"label": "Чемпионы", "color": "#22c55e", "icon": "crown",
                  "desc": "Покупают часто, много, недавно", "action": "VIP программа, эксклюзивные предложения"},
    "loyal": {"label": "Лояльные", "color": "#3b82f6", "icon": "star",
              "desc": "Регулярные, хорошие суммы", "action": "Программа лояльности, бонусы за повторные"},
    "potential_loyal": {"label": "Перспективные", "color": "#8b5cf6", "icon": "zap",
                        "desc": "Недавние, но редкие", "action": "Стимулировать повторные покупки"},
    "new_customers": {"label": "Новые", "color": "#06b6d4", "icon": "users",
                      "desc": "Только зарегистрировались", "action": "Welcome бонус, знакомство с акциями"},
    "sleeping": {"label": "Засыпающие", "color": "#f59e0b", "icon": "clock",
                 "desc": "Давно не были", "action": "Напоминание о бонусах, спец. предложение"},
    "at_risk": {"label": "Под угрозой", "color": "#ef4444", "icon": "alert-triangle",
                "desc": "Раньше были активны, исчезли", "action": "Срочная реактивация, промокод"},
    "lost": {"label": "Потерянные", "color": "#64748b", "icon": "eye-off",
             "desc": "Давно не покупали", "action": "Агрессивный win-back или исключить"},
}


@router.get("/rfm-pro", dependencies=[Depends(require_role(UserRole.SUPER_ADMIN, UserRole.BRANCH_ADMIN))])
async def rfm_pro(
    db: AsyncSession = Depends(get_db),
) -> dict:
    """RFM Pro — сегментация с рекомендациями и метриками."""
    now = datetime.now(timezone.utc)

    q = await db.execute(
        select(
            Customer.id,
            Customer.full_name,
            Customer.phone,
            Customer.created_at.label("registered"),
            func.max(Transaction.created_at).label("last_purchase"),
            func.count(Transaction.id).label("frequency"),
            func.coalesce(func.sum(Transaction.purchase_amount), 0).label("monetary"),
        ).outerjoin(Transaction, and_(
            Transaction.customer_id == Customer.id,
            Transaction.type == TransactionType.EARN,
        )).where(Customer.is_active == True)
        .group_by(Customer.id, Customer.full_name, Customer.phone, Customer.created_at)
    )
    customers = q.all()

    if not customers:
        return {"segments": {}, "total": 0, "meta": RFM_SEGMENTS}

    # Calculate percentiles
    recencies = []
    frequencies = []
    monetaries = []
    for c in customers:
        if c.last_purchase:
            recencies.append((now - c.last_purchase).days)
        frequencies.append(c.frequency)
        monetaries.append(float(c.monetary))

    def pct(arr, p):
        if not arr:
            return 0
        s = sorted(arr)
        k = (len(s) - 1) * p / 100
        f = int(k)
        ci = min(f + 1, len(s) - 1)
        return s[f] + (k - f) * (s[ci] - s[f])

    r33, r66 = (pct(recencies, 33) or 7), (pct(recencies, 66) or 30)
    f33, f66 = (pct(frequencies, 33) or 2), (pct(frequencies, 66) or 5)
    m33, m66 = (pct(monetaries, 33) or 1000), (pct(monetaries, 66) or 5000)

    seg_data: dict = {k: {"count": 0, "revenue": 0, "customers": []} for k in RFM_SEGMENTS}

    for c in customers:
        money = float(c.monetary)
        freq = c.frequency

        if not c.last_purchase:
            seg = "new_customers" if freq == 0 else "lost"
        else:
            days_ago = (now - c.last_purchase).days
            if days_ago <= r33 and freq >= f66 and money >= m66:
                seg = "champions"
            elif freq >= f66 and money >= m33:
                seg = "loyal"
            elif days_ago <= r33 and freq <= f33:
                seg = "potential_loyal"
            elif days_ago <= r33:
                seg = "new_customers"
            elif days_ago <= r66:
                seg = "sleeping"
            elif days_ago <= r66 * 2:
                seg = "at_risk"
            else:
                seg = "lost"

        seg_data[seg]["count"] += 1
        seg_data[seg]["revenue"] += money
        seg_data[seg]["customers"].append({
            "id": str(c.id),
            "name": c.full_name or "—",
            "phone": c.phone,
            "frequency": freq,
            "monetary": money,
            "last_purchase": c.last_purchase.isoformat() if c.last_purchase else None,
            "days_ago": (now - c.last_purchase).days if c.last_purchase else None,
        })

    total = len(customers)
    segments_out = {}
    for key, data in seg_data.items():
        meta = RFM_SEGMENTS[key]
        cnt = data["count"]
        # Sort customers by monetary desc, limit to top 20
        data["customers"].sort(key=lambda x: x["monetary"], reverse=True)
        segments_out[key] = {
            "count": cnt,
            "percent": round(cnt / max(total, 1) * 100, 1),
            "total_revenue": round(data["revenue"], 0),
            "avg_revenue": round(data["revenue"] / max(cnt, 1), 0),
            "label": meta["label"],
            "color": meta["color"],
            "icon": meta["icon"],
            "description": meta["desc"],
            "action": meta["action"],
            "top_customers": data["customers"][:20],
        }

    # Revenue concentration
    total_rev = sum(s["total_revenue"] for s in segments_out.values())
    for s in segments_out.values():
        s["revenue_share"] = round(s["total_revenue"] / max(total_rev, 1) * 100, 1)

    return {
        "segments": segments_out,
        "total": total,
        "total_revenue": total_rev,
        "meta": RFM_SEGMENTS,
    }
